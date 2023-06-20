
import math
import statistics
import openpyxl
import car
import aformulas as form
arr_headings_output_1 = []
arr_value_output_1 = []

arr_headings_output_2 = []
arr_value_output_2 = []

arr_headings_output_3 = []
arr_value_output_3 = []

ONTP = openpyxl.open('ONTP.xlsx', read_only=True)
table_1 = ONTP.worksheets[0]
table_2 = ONTP.worksheets[1]
table_3 = ONTP.worksheets[2]
table_4 = ONTP.worksheets[3]
table_5 = ONTP.worksheets[4]
table_6 = ONTP.worksheets[5]
table_7 = ONTP.worksheets[6]
table_8 = ONTP.worksheets[7]
table_9 = ONTP.worksheets[8]
table_10 = ONTP.worksheets[9]
table_11 = ONTP.worksheets[10]
table_12 = ONTP.worksheets[11]
table_12 = ONTP.worksheets[12]
table_13 = ONTP.worksheets[13]
table_14 = ONTP.worksheets[14]
table_10 = ONTP.worksheets[15]
table_17 = ONTP.worksheets[16]
table_19 = ONTP.worksheets[18]
table_21 = ONTP.worksheets[20]
table_24 = ONTP.worksheets[23]
table_27 = ONTP.worksheets[26]
table_29 = ONTP.worksheets[28]
table_31 = ONTP.worksheets[30]
table_32 = ONTP.worksheets[31]
table_38 = ONTP.worksheets[37]
table_2_3 = ONTP.worksheets[46]
table_2_4 = ONTP.worksheets[47]
table_2_8 = ONTP.worksheets[51]

uaz_3163 = car.Vehicle("УАЗ-3163", 29, 2.7, 20, 'автомобили легковые',
                       'автомобили и автобусы полноприводные', 122, 300000, 4.647, 1.929, 'close')
kamaz_43502 = car.Vehicle("КАМАЗ-43502", 124, 4, 105, 'автомобили грузовые',
                          'базовый автомобиль', 89, 300000, 7.570, 2.5, 'close')
kamaz_43118 = car.Vehicle("КАМАЗ-43118", 144, 10, 93,
                          'автомобили грузовые', 'седельные тягачи', 107, 300000, 8.835, 2.5, 'open')
politrans = car.Vehicle("ПОЛИТРАНС-94163", 61, 40, 32,
                        'полуприцепы тяжеловозы', 'полуприцепы', 113, 250000, 15.130, 2.550, 'open')
cars = [uaz_3163, kamaz_43502, kamaz_43118, politrans]


DAYS_JOB = 255
category = 3
weather = "холодные"
all_quantity = 0
P_proizvod_rab = 66
F_sklad_vne_proizvod_korp = 0
for i in cars:
    all_quantity += i.quantity
LEN_CAR = len(cars)

S_cars = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    S_cars[i] = round(cars[i].length * cars[i].wigth, 2)
average_S = 0
for i in range(LEN_CAR):
    average_S = average_S + S_cars[i] * cars[i].quantity
average_S = round(average_S/all_quantity, 2)

# 1.2 Lk
k1_lk = form.find_k1_for_lk(category, table_5)
k2_lk_arr = form.find_k2_for_lk(cars, table_6)
k3_lk = form.find_k3_for_lk(weather, table_7)


def Lk_car(car, k2):
    return round((((car.Lk_normat * car.new_vihicle + ((car.Lk_normat*0.8) * (car.quantity - car.new_vihicle))) / (car.quantity)))*k1_lk*k2*k3_lk, 2)


Lk_corrected = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    Lk_corrected[i] = Lk_car(cars[i], k2_lk_arr[i])

# L_to Откорректированный пробег между ТО
k1_to = form.find_k1_for_l_to(category, table_5)
k3_to = form.find_k3_for_l_to(weather, table_7)
L_to_1_normat_arr = form.find_normat_l_to1(cars, table_4)
L_to_2_normat_arr = form.find_normat_l_to2(cars, table_4)

L_to_1_cars = [i for i in range(LEN_CAR)]
L_to_2_cars = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    L_to_1_cars[i] = L_to_1_normat_arr[i] * k1_to * k3_to
    L_to_2_cars[i] = L_to_2_normat_arr[i] * k1_to * k3_to

# n_to количество ТО по дневному пробегу
n_to_1_cars = [i for i in range(LEN_CAR)]
n_to_2_cars = [i for i in range(LEN_CAR)]
L_to_1_cars_corrected = [i for i in range(LEN_CAR)]
L_to_2_cars_corrected = [i for i in range(LEN_CAR)]
chemy_kratno_n = 4
for i in range(LEN_CAR):
    n_to_1_cars[i] = round(L_to_1_cars[i]/cars[i].lcc)
    while n_to_1_cars[i] % chemy_kratno_n != 0:
        if n_to_1_cars[i] % chemy_kratno_n >= 2:
            n_to_1_cars[i] += 1
        elif n_to_1_cars[i] % chemy_kratno_n <= 2:
            n_to_1_cars[i] -= 1
    L_to_1_cars_corrected[i] = n_to_1_cars[i] * cars[i].lcc
    n_to_2_cars[i] = round(L_to_2_cars[i]/L_to_1_cars[i])
    while n_to_2_cars[i] % chemy_kratno_n != 0:
        if n_to_2_cars[i] > 4:
            if n_to_2_cars[i] % chemy_kratno_n >= 2:
                n_to_2_cars[i] += 1
            elif n_to_2_cars[i] % chemy_kratno_n <= 2:
                n_to_2_cars[i] -= 1
        else:
            n_to_2_cars[i] = 4
    L_to_2_cars_corrected[i] = L_to_1_cars_corrected[i] * n_to_2_cars[i]


def append_1_5_1(arr1, arr2):

    arr1.append(f'all_quantity')
    arr2.append(all_quantity)
    arr1.append(f'DAYS_JOB')
    arr2.append(DAYS_JOB)
    arr1.append(f'k1_lk')
    arr2.append(k1_lk)
    arr1.append(f'k3_lk')
    arr2.append(k3_lk)
    for i in range(LEN_CAR):
        arr1.append(f'lcc{i}')
        arr2.append(cars[i].lcc)
    for i in range(LEN_CAR):
        arr1.append(f'Lk_normat{i}')
        arr2.append(cars[i].Lk_normat)
        arr1.append(f'Lk_normat_before_kap{i}')
        arr2.append(round(cars[i].Lk_normat*0.8, 2))
        arr1.append(f'car.new_vihicle{i}')
        arr2.append(cars[i].new_vihicle)
        arr1.append(f'car.old_vihicle{i}')
        arr2.append(cars[i].quantity - cars[i].new_vihicle)
        arr1.append(f'k2_lk_arr{i}')
        arr2.append(k2_lk_arr[i])
        arr1.append(f'car.quantity{i}')
        arr2.append(cars[i].quantity)
        arr1.append(f'Lk_corrected{i}')
        arr2.append(Lk_corrected[i])
    arr1.append(f'k1_to')
    arr2.append(k1_to)
    arr1.append(f'k3_to')
    arr2.append(k3_to)
    for i in range(LEN_CAR):
        arr1.append(f'L_to_1_normat_arr{i}')
        arr2.append(L_to_1_normat_arr[i])
        arr1.append(f'L_to_1_cars{i}')
        arr2.append(L_to_1_cars[i])
    for i in range(LEN_CAR):
        arr1.append(f'L_to_2_normat_arr{i}')
        arr2.append(L_to_2_normat_arr[i])
        arr1.append(f'L_to_2_cars{i}')
        arr2.append(L_to_2_cars[i])
    for i in range(LEN_CAR):
        arr1.append(f'n_to_1_cars{i}')
        arr2.append(n_to_1_cars[i])
    for i in range(LEN_CAR):
        arr1.append(f'L_to_1_cars_corrected{i}')
        arr2.append(L_to_1_cars_corrected[i])
    for i in range(LEN_CAR):
        arr1.append(f'n_to_2_cars{i}')
        arr2.append(n_to_2_cars[i])
    for i in range(LEN_CAR):
        arr1.append(f'L_to_2_cars_corrected{i}')
        arr2.append(L_to_2_cars_corrected[i])


append_1_5_1(arr_headings_output_1, arr_value_output_1)

# 1.3 N_to_eo_cycle
N_1_cycle = [i for i in range(LEN_CAR)]
N_2_cycle = [i for i in range(LEN_CAR)]
N_EO_smen_cycle = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    N_2_cycle[i] = round(Lk_corrected[i] / L_to_2_cars_corrected[i] - 1, 2)
    N_1_cycle[i] = round(
        Lk_corrected[i] / L_to_1_cars_corrected[i] - 1 - N_2_cycle[i], 2)
    N_EO_smen_cycle[i] = round(Lk_corrected[i] / cars[i].lcc, 2)

# 1.5 n_g
arr_inTOTR_inKR = form.find_inTOTR_inKR(cars, table_1)
k2_TO_TR = form.find_k2_for_TO_TR(cars, table_6)
a_t = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    a_t[i] = round(1 / (1 + cars[i].lcc * (arr_inTOTR_inKR[i][0]
                   * k2_TO_TR[i] / 1000 + arr_inTOTR_inKR[i][1]/Lk_corrected[i])), 2)
Lg = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    Lg[i] = round(DAYS_JOB * cars[i].lcc * a_t[i], 2)
n_g = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    n_g[i] = round(Lg[i] / Lk_corrected[i], 2)

# 1.6 N_to_eo_year
N_1_year = [i for i in range(LEN_CAR)]
N_2_year = [i for i in range(LEN_CAR)]
N_EO_smen_year = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    N_1_year[i] = round(N_1_cycle[i] * n_g[i], 2)
    N_2_year[i] = round(N_2_cycle[i] * n_g[i], 2)
    N_EO_smen_year[i] = round(N_EO_smen_cycle[i] * n_g[i], 2)

# 1.7 N_to_eo_year_all
N_1_year_all = [i for i in range(LEN_CAR)]
N_2_year_all = [i for i in range(LEN_CAR)]
N_EO_smen_year_all = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    N_1_year_all[i] = round(N_1_year[i] * cars[i].quantity, 2)
    N_2_year_all[i] = round(N_2_year[i] * cars[i].quantity, 2)
    N_EO_smen_year_all[i] = round(N_EO_smen_year[i] * cars[i].quantity, 2)


# 1.8 N_D_1_2
N_D1_year_all = [i for i in range(LEN_CAR)]
N_D2_year_all = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    N_D1_year_all[i] = round(1.4 * N_1_year_all[i] + N_2_year_all[i], 2)
    N_D2_year_all[i] = round(1.3 * N_2_year_all[i], 2)

# 1.9 N_EO_TO_D_sutki
N_1_day = [i for i in range(LEN_CAR)]
N_2_day = [i for i in range(LEN_CAR)]
N_EO_smen_day = [i for i in range(LEN_CAR)]
N_D1_day = [i for i in range(LEN_CAR)]
N_D2_day = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    def find_day_prog(N_year_all, DAYS_JOB):
        if N_year_all/DAYS_JOB < 1:
            return 1
        else:
            return round(N_year_all / DAYS_JOB)

    N_1_day[i] = find_day_prog(N_1_year_all[i], DAYS_JOB)
    N_2_day[i] = find_day_prog(N_2_year_all[i], DAYS_JOB)
    N_EO_smen_day[i] = find_day_prog(N_EO_smen_year_all[i], DAYS_JOB)
    N_D1_day[i] = find_day_prog(N_D1_year_all[i], DAYS_JOB)
    N_D2_day[i] = find_day_prog(N_D2_year_all[i], DAYS_JOB)

N_1_day_sum = round(sum(N_1_day))
N_2_day_sum = round(sum(N_2_day))
N_EO_smen_day_sum = round(sum(N_EO_smen_day))
N_D1_day_sum = round(sum(N_D1_day))
N_D2_day_sum = round(sum(N_D2_day))

# 1.10 Metod org
metod_EO = form.metod_EO(N_EO_smen_day_sum)
metod_to1 = form.metod_to1(N_1_day_sum)
metod_to2 = form.metod_to2(N_2_day_sum)
metod_D1 = form.metod_D1(metod_to1, metod_to2)
metod_D2 = 'Метод организации на универсальных постах'


def append_1():
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_1_cycle{i}')
        arr_value_output_1.append(N_1_cycle[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_2_cycle{i}')
        arr_value_output_1.append(N_2_cycle[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_EO_smen_cycle{i}')
        arr_value_output_1.append(N_EO_smen_cycle[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'a_t{i}')
        arr_value_output_1.append(a_t[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'Lg{i}')
        arr_value_output_1.append(Lg[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'n_g{i}')
        arr_value_output_1.append(n_g[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_1_year{i}')
        arr_value_output_1.append(N_1_year[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_2_year{i}')
        arr_value_output_1.append(N_2_year[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_EO_smen_year{i}')
        arr_value_output_1.append(N_EO_smen_year[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_1_year_all{i}')
        arr_value_output_1.append(N_1_year_all[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_2_year_all{i}')
        arr_value_output_1.append(N_2_year_all[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_EO_smen_year_all{i}')
        arr_value_output_1.append(N_EO_smen_year_all[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_D1_year_all{i}')
        arr_value_output_1.append(N_D1_year_all[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_D2_year_all{i}')
        arr_value_output_1.append(N_D2_year_all[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_1_day{i}')
        arr_value_output_1.append(N_1_day[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_2_day{i}')
        arr_value_output_1.append(N_2_day[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_EO_smen_day{i}')
        arr_value_output_1.append(N_EO_smen_day[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_D1_day{i}')
        arr_value_output_1.append(N_D1_day[i])
    for i in range(LEN_CAR):
        arr_headings_output_1.append(f'N_D2_day{i}')
        arr_value_output_1.append(N_D2_day[i])
    arr_headings_output_1.append(f'N_1_day_sum')
    arr_value_output_1.append(N_1_day_sum)
    arr_headings_output_1.append(f'N_2_day_sum')
    arr_value_output_1.append(N_2_day_sum)
    arr_headings_output_1.append(f'N_EO_smen_day_sum')
    arr_value_output_1.append(N_EO_smen_day_sum)
    arr_headings_output_1.append(f'N_D1_day_sum')
    arr_value_output_1.append(N_D1_day_sum)
    arr_headings_output_1.append(f'N_D2_day_sum')
    arr_value_output_1.append(N_D2_day_sum)
    arr_headings_output_1.append(f'metod_EO')
    arr_value_output_1.append(metod_EO)
    arr_headings_output_1.append(f'metod_to1')
    arr_value_output_1.append(metod_to1)
    arr_headings_output_1.append(f'metod_to2')
    arr_value_output_1.append(metod_to2)
    arr_headings_output_1.append(f'metod_D1')
    arr_value_output_1.append(metod_D1)
    arr_headings_output_1.append(f'metod_D2')
    arr_value_output_1.append(metod_D2)


append_1()

# 2.1 trudoemkost
trud_normat = form.find_trud_normat(cars, table_1)
k1_trud = form.find_k1_for_trud_TR(category, table_5)
k2_trud = form.find_k2_for_trud(cars, table_6)
k3_trud = form.find_k3_for_trud_TR(weather, table_7)
k4_trud = form.find_k4_trud(cars, table_8)
k5_trud = form.find_k5_trud(cars, table_7)


def append_1_5_1(arr1, arr2):
    for i in range(LEN_CAR):
        for j in range(len(trud_normat[i])):
            arr1.append(f'trud_normat{i}')
            arr2.append(trud_normat[i][j])
    arr1.append(f'k1_trud')
    arr2.append(k1_trud)
    for i in range(LEN_CAR):
        arr1.append(f'k2_trud{i}')
        arr2.append(k2_trud[i])
    arr1.append(f'k3_trud')
    arr2.append(k3_trud)
    for i in range(LEN_CAR):
        arr1.append(f'k4_trud{i}')
        arr2.append(k4_trud[i])
    for i in range(LEN_CAR):
        arr1.append(f'k5_trud{i}')
        arr2.append(k5_trud[i])

append_1_5_1(arr_headings_output_1, arr_value_output_1)

t_EO = [i for i in range(LEN_CAR)]
t_to1 = [i for i in range(LEN_CAR)]
t_to2 = [i for i in range(LEN_CAR)]
t_tr1000 = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    t_EO[i] = round(trud_normat[i][0] * k2_trud[i], 2)
    t_to1[i] = round(trud_normat[i][1] * k2_trud[i] * k4_trud[i], 2)
    t_to2[i] = round(trud_normat[i][2] * k2_trud[i] * k4_trud[i], 2)
    t_tr1000[i] = round(trud_normat[i][3] * k1_trud * k2_trud[i] * k3_trud * k4_trud[i] * k5_trud[i], 2)


# 2.2 Trud_year
T_EO_year = [i for i in range(LEN_CAR)]
T_to1_year = [i for i in range(LEN_CAR)]
T_to2_year = [i for i in range(LEN_CAR)]
T_tr1000_year = [i for i in range(LEN_CAR)]
T_D1_year = [i for i in range(LEN_CAR)]
T_D2_year = [i for i in range(LEN_CAR)]
T_to1_year_all_corrected = [i for i in range(LEN_CAR)]
T_to2_year_all_corrected = [i for i in range(LEN_CAR)]
T_tr1000_year_all_corrected = [i for i in range(LEN_CAR)]

for i in range(LEN_CAR):
    T_EO_year[i] = round(t_EO[i] * N_EO_smen_year_all[i], 2)
    T_to1_year[i] = round(t_to1[i] * N_1_year_all[i], 2)
    T_to2_year[i] = round(t_to2[i] * N_2_year_all[i], 2)
    T_tr1000_year[i] = round((Lg[i] * cars[i].quantity * t_tr1000[i]) / 1000, 2)

fractions_EO = form.find_fractions(cars, table_17, T_EO_year, 3, 8)
fractions_TO_1 = form.find_fractions(cars, table_17, T_to1_year, 10, 12)
fractions_TO_2 = form.find_fractions(cars, table_17, T_to2_year, 14, 16)
fractions_post = form.find_fractions(cars, table_17, T_tr1000_year, 19, 24)
fractions_uch = form.find_fractions(cars, table_17, T_tr1000_year, 26, 40)

for i in range(LEN_CAR):

    
    T_D1_year[i] = fractions_TO_1[i][0][2] + fractions_post[i][0][2]
    T_D2_year[i] = fractions_TO_2[i][0][2] + fractions_post[i][1][2]
    T_to1_year_all_corrected[i] = T_to1_year[i] - fractions_TO_1[i][0][2]
    T_to2_year_all_corrected[i] = T_to2_year[i] -  fractions_TO_2[i][0][2] 
    T_tr1000_year_all_corrected[i] = T_tr1000_year[i] - fractions_post[i][0][2] - fractions_post[i][1][2]
    def append_Tttt(arr1, arr2):
        arr1.append(f't_D_1_coorect{i}')
        arr2.append(f"{T_to1_year[i]}⋅{round(fractions_TO_1[i][0][1]/100, 2)}+{T_tr1000_year[i]}⋅{round(fractions_post[i][0][1]/100, 2)}="+
                    f"{round(T_D1_year[i], 2)}")
        arr1.append(f't_D_2_coorect{i}')
        arr2.append(f"{T_to2_year[i]}⋅{round(fractions_TO_2[i][0][1]/100, 2)}+{T_tr1000_year[i]}⋅{round(fractions_post[i][1][1]/100, 2)}="+
                    f"{round(T_D2_year[i], 2)}")
        
        arr1.append(f't_to1_coorect{i}')
        arr2.append(f"{fractions_TO_1[i][0][2]}={round(T_to1_year_all_corrected[i], 2)}")

        arr1.append(f't_to2_coorect{i}')
        arr2.append(f"{fractions_TO_2[i][0][2]}={round(T_to2_year_all_corrected[i], 2)}")
        
        arr1.append(f't_tr_coorect{i}')
        arr2.append(f"({fractions_post[i][0][2]}+{fractions_post[i][1][2]})={round(T_tr1000_year_all_corrected[i], 2)}")
    append_Tttt(arr_headings_output_1, arr_value_output_1)
T_EO_year_all = sum(T_EO_year)
T_to1_year_all = sum(T_to1_year_all_corrected)
T_to2_year_all = sum(T_to2_year_all_corrected)
T_tr1000_year_all = sum(T_tr1000_year_all_corrected)
T_D1_year_all = sum(T_D1_year)
T_D2_year_all = sum(T_D2_year)
T_vsp = (T_EO_year_all + T_to1_year_all + T_to2_year_all + T_tr1000_year_all) * 0.2


def append_2(arr1, arr2):
    for i in range(LEN_CAR):
        arr1.append(f't_EOc{i}')
        arr2.append(t_EO[i])
    for i in range(LEN_CAR):
        arr1.append(f't_to1{i}')
        arr2.append(t_to1[i])
    for i in range(LEN_CAR):
        arr1.append(f't_to2{i}')
        arr2.append(t_to2[i])
    for i in range(LEN_CAR):
        arr1.append(f't_tr1000{i}')
        arr2.append(t_tr1000[i])


    for i in range(LEN_CAR):
        arr1.append(f'T_EO_year{i}')
        arr2.append(T_EO_year[i])


    for i in range(LEN_CAR):
        arr1.append(f'T_to1_year{i}')
        arr2.append(T_to1_year[i])
    for i in range(LEN_CAR):
        arr1.append(f'T_to2_year{i}')
        arr2.append(T_to2_year[i])
    for i in range(LEN_CAR):
        arr1.append(f'T_tr1000_year{i}')
        arr2.append(T_tr1000_year[i])
    arr1.append(f'T_EO_year_all')
    arr2.append(T_EO_year_all)
    arr1.append(f'T_D1_year_all')
    arr2.append(T_D1_year_all)
    arr1.append(f'T_D2_year_all')
    arr2.append(T_D2_year_all)
    arr1.append(f'T_to1_year_all')
    arr2.append(T_to1_year_all)
    arr1.append(f'T_to2_year_all')
    arr2.append(T_to2_year_all)
    arr1.append(f'T_tr1000_year_all')
    arr2.append(T_tr1000_year_all)
    arr1.append(f'T_vsp')
    arr2.append(T_vsp)
append_2(arr_headings_output_1, arr_value_output_1)


# table 1.1 and 1.2



def find_trud_post_and_uch(fractions):
    tmp_trud = 0
    for i in range(LEN_CAR):
        for j in range(len(fractions[i])):
            tmp_trud = tmp_trud + fractions[i][j][2]
    return round(tmp_trud, 2)


trud_post_all = find_trud_post_and_uch(fractions_post)
trud_uch_all = find_trud_post_and_uch(fractions_uch)
count_row = 2

modes_to_1 = form.days_job_for_to1(DAYS_JOB)
modes_to_2 = form.days_job_for_to2(DAYS_JOB)


# №№№№№№№№№№№№№№№№№№№№№ Часть ДВА
# 3 X TO-1 X TO-2 X D
X_to_1_arr = [i for i in range(LEN_CAR)]
X_to_2_arr = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    def appendx_to(arr1, arr2):
        arr1.append(f'N_1_day{i}')
        arr2.append(N_1_day[i])
        arr1.append(f't_to1_corrected{i}')
        arr2.append(t_to1[i])
        arr1.append(f'ritm_to_1{i}')
        arr2.append(ritm_to_1)
        arr1.append(f'takt_to_1{i}')
        arr2.append(takt_to_1)
        arr1.append(f'X_to_1{i}')
        arr2.append(X_to_1)
        arr1.append(f'N_2_day{i}')
        arr2.append(N_2_day[i])
        arr1.append(f'ritm_to_2{i}')
        arr2.append(ritm_to_2)
        arr1.append(f't_to2_corrected{i}')
        arr2.append(t_to2[i])
        arr1.append(f'takt_to_2{i}')
        arr2.append(takt_to_2)
        arr1.append(f'X_to_2{i}')
        arr2.append(X_to_2)
    ritm_to_1 = round((60*modes_to_1[2]*modes_to_1[1])/N_1_day[i], 2)
    R_p_to_1 = 2
    time_peredviz_on_to = 2
    takt_to_1 = round((60*t_to1[i]/R_p_to_1)+time_peredviz_on_to, 2)
    X_to_1 = round(takt_to_1/ritm_to_1)
    if X_to_1 == 0:
        X_to_1 = 1
    if cars[i].modif_vihicle == 'полуприцепы':
        X_to_1 = 0
    X_to_1_arr[i] = X_to_1

    ritm_to_2 = round((60*modes_to_2[2]*modes_to_2[1])/N_2_day[i], 2)
    R_p_to_2 = 3
    n_to_2 = 0.98
    takt_to_2 = round((60*t_to2[i]/R_p_to_2)+time_peredviz_on_to, 2)
    X_to_2 = round(takt_to_2/(ritm_to_2*n_to_2))
    if X_to_2 == 0:
        X_to_2 = 1
    if cars[i].modif_vihicle == 'полуприцепы':
        X_to_2 = 0
    X_to_2_arr[i] = X_to_2
 
    if cars[i].modif_vihicle != 'полуприцепы':
        appendx_to(arr_headings_output_2, arr_value_output_2)
X_to_1 = sum(X_to_1_arr)
X_to_2 = sum(X_to_2_arr)

def append3(arr1, arr2):
    for i in range(LEN_CAR):
        arr1.append(f'cars[i].name')
        arr2.append(cars[i].name)
    arr1.append(f'T_sm_to1')
    arr2.append(modes_to_1[2])
    arr1.append(f'C_to1')
    arr2.append(modes_to_1[1])
    arr1.append(f'N_1_day_sum')
    arr2.append(N_1_day_sum)
    arr1.append(f'R_p_to_1')
    arr2.append(R_p_to_1)
    arr1.append(f'time_peredviz_on_to')
    arr2.append(time_peredviz_on_to)
    arr1.append(f'T_cm_to_1')
    arr2.append(modes_to_1[2])
    arr1.append(f'C_to_1')
    arr2.append(modes_to_1[1])
    arr1.append(f'N_2_day_sum')
    arr2.append(N_2_day_sum)
    arr1.append(f'R_p_to_2')
    arr2.append(R_p_to_2)
    arr1.append(f'n_to_2')
    arr2.append(n_to_2)
    arr1.append(f'T_cm_to_2')
    arr2.append(modes_to_2[2])
    arr1.append(f'C_to_2')
    arr2.append(modes_to_2[1])
    arr1.append(f'X_to_1')
    arr2.append(X_to_1)
    arr1.append(f'X_to_2')
    arr2.append(X_to_2)
append3(arr_headings_output_2, arr_value_output_2)


# D
modes_D = form.days_job_for_D(DAYS_JOB)
n_d = 0.75
R_p_D1 = 2
R_p_D2 = 1
X_d_1 = round(T_D1_year_all / (DAYS_JOB *
              modes_D[1] * modes_D[2] * n_d * R_p_D1))
X_d_2 = round(T_D2_year_all / (DAYS_JOB *
              modes_D[1] * modes_D[2] * n_d * R_p_D2))
if X_d_1 == 0:
    X_d_1 = 1
if X_d_2 == 0:
    X_d_2 = 1

def append_3_2_15(arr1, arr2):
    arr1.append(f'T_D1_year_all')
    arr2.append(T_D1_year_all)
    arr1.append(f'T_D2_year_all')
    arr2.append(T_D2_year_all)
    arr1.append(f'DAYS_JOB')
    arr2.append(DAYS_JOB)
    arr1.append(f'T_sm_on_D')
    arr2.append(modes_D[2])
    arr1.append(f'C_on_D')
    arr2.append(modes_D[1])
    arr1.append(f'n_d')
    arr2.append(n_d)
    arr1.append(f'R_p_D1')
    arr2.append(R_p_D1)
    arr1.append(f'R_p_D2')
    arr2.append(R_p_D2)
    arr1.append(f'X_d_1')
    arr2.append(X_d_1)
    arr1.append(f'X_d_2')
    arr2.append(X_d_2)
append_3_2_15(arr_headings_output_2, arr_value_output_2)

# EO
N_y_for_EO = 40
takt_EO_line = round(60/N_y_for_EO, 2)
N_EO_line = round(60/takt_EO_line, 2)
T_voz_EO = 3.3
R_EO_line = round(60 * (T_voz_EO/(0.7 * N_EO_smen_day_sum)), 2)
R_p_EO_line = 2
m_EO_line = round(takt_EO_line/R_EO_line)
if m_EO_line == 0:
    m_EO_line = 1
modes_EO = form.days_job_for_EO(DAYS_JOB, table_27)
n_2_for_EO = form.find_n_2_for_EO(modes_EO[1], table_29)
fi_for_EO = form.find_fi('EO', all_quantity, modes_EO[1], table_24)
X_EO_line = round((T_EO_year_all*fi_for_EO) /
                  (DAYS_JOB*modes_EO[2]*n_2_for_EO[0]*modes_EO[1]*R_p_EO_line))
if X_EO_line == 0:
    X_EO_line = 1


def append_3_2_16(arr1, arr2):
    arr1.append(f'N_EO_line')
    arr2.append(N_EO_line)
    arr1.append(f'N_y_for_EO')
    arr2.append(N_y_for_EO)
    arr1.append(f'takt_EO_line')
    arr2.append(takt_EO_line)
    arr1.append(f'R_EO_line')
    arr2.append(R_EO_line)
    arr1.append(f'm_EO_line')
    arr2.append(m_EO_line)
    arr1.append(f'T_voz_EO')
    arr2.append(T_voz_EO)
    arr1.append(f'T_EO_year_all')
    arr2.append(T_EO_year_all)
    arr1.append(f'n_2_for_EO[0]')
    arr2.append(n_2_for_EO[0])
    arr1.append(f'X_EO_line')
    arr2.append(X_EO_line)
    arr1.append(f'fi_for_EO')
    arr2.append(fi_for_EO)
    arr1.append(f'modes_EO[2]')
    arr2.append(modes_EO[2])
    arr1.append(f'modes_EO[1]')
    arr2.append(modes_EO[1])
    arr1.append(f'R_p_EO_line')
    arr2.append(R_p_EO_line)
append_3_2_16(arr_headings_output_2, arr_value_output_2)

# TR
modes_TR = form.days_job_for_TR(DAYS_JOB, table_27)

R_p_TR = 2
n_2_for_TR = form.find_n_2_for_TR(modes_TR[1], table_29)

def find_trud_post_one_car(fractions):
    trud = []
    for i in range(LEN_CAR):
        tmp_trud = 0
        for j in range(len(fractions[i])):
            tmp_trud = tmp_trud + fractions[i][j][2]
        trud.append(round(tmp_trud, 2))
    return trud
trud_post_unique_car = find_trud_post_one_car(fractions_post)
X_tr_arr = [i for i in range(LEN_CAR)]
for i in range(LEN_CAR):
    fi_for_TR = form.find_fi('TR', cars[i].quantity, modes_TR[1], table_24)
    
    X_tr_ne_okrug = round((trud_post_unique_car[i] * fi_for_TR) /
             (DAYS_JOB*modes_TR[2]*R_p_TR*n_2_for_TR*modes_TR[1]), 2)
    X_tr = round((trud_post_unique_car[i] * fi_for_TR) /
             (DAYS_JOB*modes_TR[2]*R_p_TR*n_2_for_TR*modes_TR[1]))

    if X_tr == 0:
        X_tr = 1
    X_tr_arr[i] = X_tr
    arr_headings_output_2.append(f'fi_for_TR{i}')
    arr_value_output_2.append(fi_for_TR)
    arr_headings_output_2.append(f'trud_post_unique_car{i}')
    arr_value_output_2.append(trud_post_unique_car[i])
    arr_headings_output_2.append(f'X_tr_ne_okrug{i}')
    arr_value_output_2.append(X_tr_ne_okrug)
    arr_headings_output_2.append(f'X_tr{i}')
    arr_value_output_2.append(X_tr)
X_tr = sum(X_tr_arr)
def append_3_2_17(arr1, arr2):
    arr_headings_output_2.append(f'trud_post_all')
    arr_value_output_2.append(trud_post_all)

    arr1.append(f'R_p_TR')
    arr2.append(R_p_TR)
    arr1.append(f'n_2_for_TR')
    arr2.append(n_2_for_TR)
    arr1.append(f'C_tr')
    arr2.append(modes_TR[1])
    arr1.append(f'T_cm_tr')
    arr2.append(modes_TR[2])
    arr1.append(f'X_tr')
    arr2.append(X_tr)
append_3_2_17(arr_headings_output_2, arr_value_output_2)

# Расчет числа постов ожидания
waiting_post_TO_1 = round(0.15 * N_1_day_sum)
waiting_post_TO_2 = round(0.35 * N_2_day_sum)
waiting_post_TR = round(0.25 * X_tr)
if waiting_post_TR == 0:
    waiting_post_TR = 1
def append_3_2_18(arr1, arr2):
    arr1.append(f'waiting_post_TO_1')
    arr2.append(waiting_post_TO_1)
    arr1.append(f'waiting_post_TO_2')
    arr2.append(waiting_post_TO_2)
    arr1.append(f'waiting_post_TR')
    arr2.append(waiting_post_TR)
append_3_2_18(arr_headings_output_2, arr_value_output_2)
#F_TO
F_z_to_1_arr = [i for i in range(LEN_CAR)]
F_z_to_2_arr = [i for i in range(LEN_CAR)]
F_z_tr_arr = [i for i in range(LEN_CAR)]
k_p_to_1 = 5
k_p_to_2 = 5
k_p_d_tr = 5
for i in range(LEN_CAR):
    def appendf_to(arr1, arr2):
        arr1.append(f'S_cars{i}')
        arr2.append(S_cars[i])
        arr1.append(f'F_z_to_1_arr{i}')
        arr2.append(F_z_to_1_arr[i])
        arr1.append(f'F_z_to_2_arr{i}')
        arr2.append(F_z_to_2_arr[i])
        arr1.append(f'F_z_tr_arr{i}')
        arr2.append(F_z_tr_arr[i])
    F_z_to_1_arr[i] = round(X_to_1_arr[i] * S_cars[i] * k_p_to_1)
    F_z_to_2_arr[i] = round(X_to_2_arr[i] * S_cars[i] * k_p_to_2)
    F_z_tr_arr[i] = round(X_tr_arr[i] * S_cars[i] * k_p_d_tr)
    appendf_to(arr_headings_output_2, arr_value_output_2)
F_z_to_1 = sum(F_z_to_1_arr)
F_z_to_2 = sum(F_z_to_2_arr)
F_z_TR = sum(F_z_tr_arr)
#F_D
X_d = X_d_1 + X_d_2
k_p_EO = 4
F_z_D = X_d * average_S * k_p_d_tr
F_z_EO = X_EO_line * average_S * k_p_EO

def okruglenie_do_krat_6(F_z):
    tmp = round(F_z/6)
    result = tmp*6
    return result
F_z_to_1_okrug = okruglenie_do_krat_6(F_z_to_1)
F_z_to_2_okrug = okruglenie_do_krat_6(F_z_to_2)
F_z_D_okrug = okruglenie_do_krat_6(F_z_D)
F_z_TR_okrug = okruglenie_do_krat_6(F_z_TR)
F_z_EO_okrug = okruglenie_do_krat_6(F_z_EO)


def append_3_2_20(arr1, arr2):

    arr1.append(f'k_p_to_1')
    arr2.append(k_p_to_1)
    arr1.append(f'k_p_to_2')
    arr2.append(k_p_to_2)

    arr1.append(f'X_d')
    arr2.append(X_d)
    arr1.append(f'k_p_d_tr')
    arr2.append(k_p_d_tr)
    arr1.append(f'S_vihicle')
    arr2.append(average_S)
    arr1.append(f'k_p_EO')
    arr2.append(k_p_EO)

    arr1.append(f'F_z_to_1')
    arr2.append(F_z_to_1)
    arr1.append(f'F_z_to_2')
    arr2.append(F_z_to_2)
    arr1.append(f'F_z_D')
    arr2.append(F_z_D)
    arr1.append(f'F_z_TR')
    arr2.append(F_z_TR)
    arr1.append(f'F_z_EO')
    arr2.append(F_z_EO)

    arr1.append(f'F_z_to_1_okrug')
    arr2.append(F_z_to_1_okrug)
    arr1.append(f'F_z_to_2_okrug')
    arr2.append(F_z_to_2_okrug)
    arr1.append(f'F_z_D_okrug')
    arr2.append(F_z_D_okrug)
    arr1.append(f'F_z_TR_okrug')
    arr2.append(F_z_TR_okrug)
    arr1.append(f'F_z_EO_okrug')
    arr2.append(F_z_EO_okrug)


append_3_2_20(arr_headings_output_2, arr_value_output_2)

# S участков
f_1_f_2_for_regions = []
f_1_f_2_for_regions = form.find_f_1_f_2_for_regions(table_31)
F_reg_all = []

# Расчет площадей складских помещений
all_sum_sklads = 0

# Расчет площади стоянки
F_x = [0 for i in range(LEN_CAR)]

for i in range(LEN_CAR):
    A_m_hra = cars[i].quantity - X_to_1_arr[i] - X_to_2_arr[i] - X_tr_arr[i]
    k_pp_or_K_for_F_x = 0
    if cars[i].metod_keeping == "close":
        k_pp_or_K_for_F_x = 2
        F_x[i] = round(S_cars[i] * A_m_hra * k_pp_or_K_for_F_x, 2)
    else:
        if cars[i].modif_vihicle == "седельные тягачи":

            k_pp_or_K_for_F_x = 1.4
            S_tmp = round((cars[i].length + cars[i+1].length) * cars[i].wigth, 2)
            F_x[i] = round(S_tmp * A_m_hra * k_pp_or_K_for_F_x, 2)
            arr_headings_output_2.append(f'S_tmp{i}')
            arr_value_output_2.append(S_tmp)
        else:
            continue
    arr_headings_output_2.append(f'A_m_hra{i}')
    arr_value_output_2.append(A_m_hra)
    arr_headings_output_2.append(f'k_pp_or_K_for_F_x{i}')
    arr_value_output_2.append(k_pp_or_K_for_F_x)

F_x_all = sum(F_x)


def append_3_2_21(arr1, arr2):
    for i in range(LEN_CAR):
        arr1.append(f'S_cars{i}')
        arr2.append(S_cars[i])
        arr1.append(f'quantity{i}')
        arr2.append(cars[i].quantity)
        arr1.append(f'F_x{i}')
        arr2.append(F_x[i])
    arr1.append(f'F_x_all')
    arr2.append(F_x_all)
append_3_2_21(arr_headings_output_2, arr_value_output_2)


def table_1_1(tables):
    def add(T_year_all, n_row, f_t, f_s):
        tables.cell(row=n_row, column=16).value = T_year_all
        tables.cell(row=n_row, column=17).value = f_t
        tables.cell(row=n_row, column=18).value = round(T_year_all / f_t, 2)
        tables.cell(row=n_row, column=19).value = f_s
        tables.cell(row=n_row, column=20).value = round(T_year_all / f_s, 2)

    def append(fractions):
        global count_row
        trud_all = 0
        for i in range(LEN_CAR):
            procent = 0
            trud = 0
            for j in range(len(fractions[i])):
                tables.cell(row=j + count_row, column=(1)
                            ).value = fractions[i][j][0]
                tables.cell(row=j + count_row, column=(2 + i*3)
                            ).value = fractions[i][j][1]
                tables.cell(row=j + count_row, column=(3 + i*3)
                            ).value = fractions[i][j][2]
                trud_one_uch = 0
                if fractions == fractions_uch and i == LEN_CAR - 1:
                    trud_one_uch = 0
                    for l in range(LEN_CAR):
                        trud_one_uch = trud_one_uch + fractions[l][j][2]
                    if fractions[i][j][0] != 'Малярные работы':
                        add(trud_one_uch, j + count_row, 2070, 1820)
                    else:
                        add(trud_one_uch, j + count_row, 1830, 1610)
                procent = procent + fractions[i][j][1]
                trud = trud + fractions[i][j][2]
            tables.cell(
                row=(count_row + len(fractions[i])), column=(2 + i*3)).value = procent
            tables.cell(
                row=(count_row + len(fractions[i])), column=(3 + i*3)).value = trud
            trud_all = trud_all + trud
        add(trud_all, count_row + len(fractions[i]), 2070, 1820)
        count_row = count_row + len(fractions[i]) + 1

    append(fractions_EO)
    append(fractions_TO_1)
    append(fractions_TO_2)
    append(fractions_post)
    append(fractions_uch)
    add(T_D1_year_all, 9, 2070, 1820)
    add(T_D2_year_all, 11, 2070, 1820)
    add(T_D2_year_all, 11, 2070, 1820)
    add(T_vsp, 35, 2070, 1820)
    T_vsp
def table_1_5(tables):
    tables.cell(row=38, column=1).value = "table_1_5"
    for i in range(len(f_1_f_2_for_regions)):
        if f_1_f_2_for_regions[i][4] != 0:
            F_reg = f_1_f_2_for_regions[i][2] + \
                f_1_f_2_for_regions[i][3] * (f_1_f_2_for_regions[i][1] - 1)
            F_reg_all.append(F_reg)
            tables.cell(row=i+39, column=1).value = f_1_f_2_for_regions[i][0]
            tables.cell(row=i+39, column=2).value = f_1_f_2_for_regions[i][2]
            tables.cell(row=i+39, column=3).value = f_1_f_2_for_regions[i][3]
            tables.cell(row=i+39, column=4).value = f_1_f_2_for_regions[i][1]
            tables.cell(row=i+39, column=5).value = F_reg
def table_1_6(tables):
    global all_sum_sklads
    for i in range(LEN_CAR):
        f_sklad_k_4 = form.find_f_sklad_k_4(cars[i].type_vihicle, table_38)
        k_1_sklad = form.find_k1_sklad(cars[i].lcc)
        k_2_sklad = form.find_k2_sklad(cars[i].quantity, table_8)
        k_3_sklad = form.find_k3_sklad(cars[i], table_1)
        k_5_sklad = form.find_k5_sklad(category, table_5)
        F_sklad_all = []

        tables.cell(row=101, column=1).value = "table_1_6"
        tmp = ""

        for j in range(len(f_sklad_k_4)):
            F_sklad = round(0.1 * cars[i].quantity * f_sklad_k_4[j][1] *
                            k_1_sklad * k_2_sklad * k_3_sklad * f_sklad_k_4[j][2] * k_5_sklad)
            if F_sklad == 0:
                F_sklad = 1
            if f_sklad_k_4[j][0] != "Кислород и ацетилен в баллонах" and f_sklad_k_4[j][0] != "Подлежащие списанию автомобили, агрегаты":
                F_sklad_all.append(F_sklad)
                if j == len(f_sklad_k_4) - 1:
                    tmp = tmp + f"{F_sklad} = {sum(F_sklad_all)}"
                    all_sum_sklads = all_sum_sklads + sum(F_sklad_all)
                else:
                    tmp = tmp + f"{F_sklad} + "
            else:
                global F_sklad_vne_proizvod_korp
                F_sklad_vne_proizvod_korp = F_sklad_vne_proizvod_korp + F_sklad
            tables.cell(row=102+i, column=1).value = cars[i].name
            tables.cell(row=101, column=2+j).value = f_sklad_k_4[j][0]
            tables.cell(
                row=102+i, column=2+j).value = f"0.1⋅{cars[i].quantity}⋅{f_sklad_k_4[j][1]}⋅{k_1_sklad}⋅{k_2_sklad}⋅{k_3_sklad}⋅{f_sklad_k_4[j][2]}⋅{k_5_sklad} = {F_sklad}"
            # tables.cell(row=102+j, column=1).value = f_sklad_k_4[j][0]
            # tables.cell(row=102+j, column=2+i).value = f"0.1⋅{cars[i].quantity}⋅{f_sklad_k_4[j][1]}⋅{k_1_sklad}⋅{k_2_sklad}⋅{k_3_sklad}⋅{f_sklad_k_4[j][2]}⋅{k_5_sklad}={F_sklad}"

        tables.cell(row=107+i, column=1).value = tmp
    arr_headings_output_2.append("all_sum_sklads")
    arr_value_output_2.append(all_sum_sklads)
def append_tables():
    book = openpyxl.Workbook()
    tables = book.active
    table_1_1(tables)
    table_1_5(tables)
    table_1_6(tables)
    book.save('tables.xlsx')
    book.close()
append_tables()


# 4.1 Расчет площадей производственного корпуса, административно-бытового корпуса и контрольно-технического пункта
F_z_all = F_z_to_1_okrug + F_z_to_2_okrug + \
    F_z_D_okrug + F_z_EO_okrug + F_z_TR_okrug
F_vsp = round(0.12 * (F_z_all + sum(F_reg_all) + all_sum_sklads))
F_proezd = round(0.1 * (F_z_all + sum(F_reg_all) + all_sum_sklads))
F_proiz_korp = F_z_all + sum(F_reg_all) + F_vsp + all_sum_sklads + F_proezd



####ЧАСТЬ ТРИИИ
A_p = 60
X_kpp = round((0.7 * all_quantity * max(a_t)) / (T_voz_EO * A_p))
if X_kpp == 0:
    X_kpp = 1
F_kpp = round(4 * X_kpp * average_S)


def append_4_1(arr1, arr2):
    arr1.append(f'all_quantity')
    arr2.append(all_quantity)
    arr1.append(f'max(a_t)')
    arr2.append(max(a_t))
    arr1.append(f'F_reg_all')
    arr2.append(sum(F_reg_all))
    arr1.append(f'F_z_all')
    arr2.append(F_z_all)
    arr1.append(f'F_vsp')
    arr2.append(F_vsp)
    arr1.append(f'F_proezd')
    arr2.append(F_proezd)
    arr1.append(f'F_proiz_korp')
    arr2.append(F_proiz_korp)
    arr1.append(f'A_p')
    arr2.append(A_p)
    arr1.append(f'X_kpp')
    arr2.append(X_kpp)
    arr1.append(f'F_kpp')
    arr2.append(F_kpp)


append_4_1(arr_headings_output_2, arr_value_output_2)

arr_headings_output_3.append(f'max(a_t)')
arr_value_output_3.append(max(a_t))
arr_headings_output_3.append(f'T_voz_EO')
arr_value_output_3.append(T_voz_EO)
arr_headings_output_3.append(f'A_p')
arr_value_output_3.append(A_p)
arr_headings_output_3.append(f'average_S')
arr_value_output_3.append(average_S)
arr_headings_output_3.append(f'X_kpp')
arr_value_output_3.append(X_kpp)
arr_headings_output_3.append(f'F_kpp')
arr_value_output_3.append(F_kpp)


# 5 ТЕХНИКО-ЭКОНОМИЧЕСКОЕ ОБОСНОВАНИЕ ПРОЕКТНЫХ РЕШЕНИЙ

X_all = X_to_1 + X_to_2 + X_tr + X_EO_line

F_abk = 0
F_gen_plan = F_proiz_korp + F_x_all + F_kpp
P_ud_etalon_all = []
X_ud_etalon_all = []
S_prsk_etalon_all = []
S_vsp_etalon_all = []
S_st_etalon_all = []
S_t_etalon_all = []
k_6_standart = form.find_k6_standart(category, table_5)
k_7_standart = form.find_k7_standart(weather, table_7)
if weather == 'холодные' or weather == 'очень холодные':
    k_5_standart = 1.27
else:
    k_5_standart = 1.16
for i in range(LEN_CAR):
    standart = form.find_standart(cars[i].type_vihicle, table_2_8)
    if len(standart) == 0:
        break
    k_1_standart = form.find_k_1_standart(cars[i].quantity, table_8)
    k_2_standart = form.find_k2_standart(cars[i], table_1)
    if cars[i].modif_vihicle == "седельные тягачи":
        procent_pricep = 0
        for k in range(LEN_CAR):
            if cars[k].modif_vihicle == 'полуприцепы':
                procent_pricep = round(cars[k].quantity/cars[i].quantity, 2)
                break
        k_3_standart = form.find_k3_standart(procent_pricep, table_2_3)

    else:
        k_3_standart = [1, 1, 1, 1, 1, 1]

    k_4_standart = form.find_k4_standart(cars[i].lcc, table_2_4)

    P_ud_etalon = round(standart[0] * k_1_standart[0] * k_2_standart[0] *
                        k_3_standart[0] * k_4_standart[0] * k_6_standart[0] * k_7_standart[0], 2)
    X_ud_etalon = round(standart[1] * k_1_standart[1] * k_2_standart[1] *
                        k_3_standart[1] * k_4_standart[1] * k_6_standart[1] * k_7_standart[1], 2)
    S_prsk_etalon = round(standart[2] * k_1_standart[2] * k_2_standart[2] *
                          k_3_standart[2] * k_4_standart[2] * k_6_standart[2] * k_7_standart[2], 2)
    S_vsp_etalon = round(standart[3] * k_1_standart[3] * k_2_standart[3] *
                         k_3_standart[3] * k_4_standart[3] * k_6_standart[3] * k_7_standart[3], 2)
    S_st_etalon = round(standart[4] * k_2_standart[4]
                        * k_3_standart[4] * k_5_standart, 2)
    S_t_etalon = round(standart[5] * k_1_standart[4] * k_2_standart[5] *
                       k_3_standart[5] * k_4_standart[4] * k_6_standart[4] * k_7_standart[4], 2)

    P_ud_etalon_all.append(P_ud_etalon)
    X_ud_etalon_all.append(X_ud_etalon)
    S_prsk_etalon_all.append(S_prsk_etalon)
    S_vsp_etalon_all.append(S_vsp_etalon)
    S_st_etalon_all.append(S_st_etalon)
    S_t_etalon_all.append(S_t_etalon)

    for j in range(len(standart)):
        arr_headings_output_3.append(f'standart{j}{i}')
        arr_value_output_3.append(standart[j])
    for j in range(len(k_1_standart)):
        arr_headings_output_3.append(f'k_1_standart{j}{i}')
        arr_value_output_3.append(k_1_standart[j])
    for j in range(len(k_2_standart)):
        arr_headings_output_3.append(f'k_2_standart{j}{i}')
        arr_value_output_3.append(k_2_standart[j])
    for j in range(len(k_3_standart)):
        arr_headings_output_3.append(f'k_3_standart{j}{i}')
        arr_value_output_3.append(k_3_standart[j])
    for j in range(len(k_4_standart)):
        arr_headings_output_3.append(f'k_4_standart{j}{i}')
        arr_value_output_3.append(k_4_standart[j])
    for j in range(len(k_6_standart)):
        arr_headings_output_3.append(f'k_6_standart{j}{i}')
        arr_value_output_3.append(k_6_standart[j])
    for j in range(len(k_7_standart)):
        arr_headings_output_3.append(f'k_7_standart{j}{i}')
        arr_value_output_3.append(k_7_standart[j])
    arr_headings_output_3.append(f'P_ud_etalon{i}')
    arr_value_output_3.append(P_ud_etalon)
    arr_headings_output_3.append(f'X_ud_etalon{i}')
    arr_value_output_3.append(X_ud_etalon)
    arr_headings_output_3.append(f'S_prsk_etalon{i}')
    arr_value_output_3.append(S_prsk_etalon)
    arr_headings_output_3.append(f'S_vsp_etalon{i}')
    arr_value_output_3.append(S_vsp_etalon)
    arr_headings_output_3.append(f'S_st_etalon{i}')
    arr_value_output_3.append(S_st_etalon)
    arr_headings_output_3.append(f'S_t_etalon{i}')
    arr_value_output_3.append(S_t_etalon)

P_ud_ne_correct = round(P_proizvod_rab/all_quantity, 2)
X_ud_ne_correct = round(X_all/all_quantity, 2)
S_pr_skl_ne_correct = round(
    (F_proiz_korp+ all_sum_sklads + F_sklad_vne_proizvod_korp)/all_quantity, 2)

S_st_ne_correct = round(sum(F_x)/all_quantity, 2)


P_ud_srav = round((P_ud_ne_correct - statistics.mean(P_ud_etalon_all))
                  * 100 / statistics.mean(P_ud_etalon_all), 2)
X_ud_srav = round((X_ud_ne_correct - statistics.mean(X_ud_etalon_all))
                  * 100 / statistics.mean(X_ud_etalon_all), 2)
S_prsk_srav = round((S_pr_skl_ne_correct - statistics.mean(S_prsk_etalon_all))
                    * 100 / statistics.mean(S_prsk_etalon_all), 2)

S_st_srav = round((S_st_ne_correct - statistics.mean(S_st_etalon_all))
                  * 100 / statistics.mean(S_st_etalon_all), 2)


F_abk = round(statistics.mean(S_vsp_etalon_all) * all_quantity/2, 2)
F_gen_plan = round((F_proiz_korp + F_abk + F_x_all + F_kpp) / 0.45, 2)
S_t_ne_correct = round(F_gen_plan/all_quantity, 2)
S_t_srav = round((S_t_ne_correct - statistics.mean(S_t_etalon_all))
                 * 100 / statistics.mean(S_t_etalon_all), 2)

S_vsp_ne_correct = round((F_vsp+F_abk+F_kpp)/all_quantity, 2)
S_vsp_srav = round((S_vsp_ne_correct - statistics.mean(S_vsp_etalon_all))
                   * 100 / statistics.mean(S_vsp_etalon_all), 2)


arr_headings_output_3.append(f'k_5_standart')
arr_value_output_3.append(k_5_standart)
arr_headings_output_3.append(f'P_ud_ne_correct')
arr_value_output_3.append(P_ud_ne_correct)
arr_headings_output_3.append(f'X_ud_ne_correct')
arr_value_output_3.append(X_ud_ne_correct)
arr_headings_output_3.append(f'S_pr_skl_ne_correct')
arr_value_output_3.append(S_pr_skl_ne_correct)
arr_headings_output_3.append(f'S_vsp_ne_correct')
arr_value_output_3.append(S_vsp_ne_correct)
arr_headings_output_3.append(f'S_st_ne_correct')
arr_value_output_3.append(S_st_ne_correct)
arr_headings_output_3.append(f'S_t_ne_correct')
arr_value_output_3.append(S_t_ne_correct)

arr_headings_output_3.append(f'P_ud_etalon_all')
arr_value_output_3.append(round(statistics.mean(P_ud_etalon_all), 2))
arr_headings_output_3.append(f'X_ud_etalon_all')
arr_value_output_3.append(round(statistics.mean(X_ud_etalon_all), 2))
arr_headings_output_3.append(f'X_ud_etalon_all')
arr_value_output_3.append(round(statistics.mean(X_ud_etalon_all), 2))
arr_headings_output_3.append(f'S_prsk_etalon_all')
arr_value_output_3.append(round(statistics.mean(S_prsk_etalon_all), 2))
arr_headings_output_3.append(f'S_vsp_etalon_all')
arr_value_output_3.append(round(statistics.mean(S_vsp_etalon_all), 2))
arr_headings_output_3.append(f'S_st_etalon_all')
arr_value_output_3.append(round(statistics.mean(S_st_etalon_all), 2))
arr_headings_output_3.append(f'S_t_etalon_all')
arr_value_output_3.append(round(statistics.mean(S_t_etalon_all), 2))


arr_headings_output_3.append(f'P_ud_srav')
arr_value_output_3.append(P_ud_srav)
arr_headings_output_3.append(f'X_ud_srav')
arr_value_output_3.append(X_ud_srav)
arr_headings_output_3.append(f'S_prsk_srav')
arr_value_output_3.append(S_prsk_srav)
arr_headings_output_3.append(f'S_vsp_srav')
arr_value_output_3.append(S_vsp_srav)
arr_headings_output_3.append(f'S_st_srav')
arr_value_output_3.append(S_st_srav)
arr_headings_output_3.append(f'S_t_srav')
arr_value_output_3.append(S_t_srav)
for i in range(LEN_CAR):
    arr_headings_output_3.append(f'cars{i}')
    arr_value_output_3.append(cars[i].name)

arr_headings_output_3.append(f'P_proizvod_rab')
arr_value_output_3.append(P_proizvod_rab)
arr_headings_output_3.append(f'X_all')
arr_value_output_3.append(X_all)
arr_headings_output_3.append(f'F_proiz_korp+F_sklad_vne_proizvod_korp')
arr_value_output_3.append(F_proiz_korp+all_sum_sklads+F_sklad_vne_proizvod_korp)
arr_headings_output_3.append(f'F_vsp+F_abk+F_kpp')
arr_value_output_3.append(F_vsp+F_abk+F_kpp)
arr_headings_output_3.append(f'sum(F_x)')
arr_value_output_3.append(F_x_all)
arr_headings_output_3.append(f'F_gen_plan')
arr_value_output_3.append(F_gen_plan)
arr_headings_output_3.append(f'all_quantity')
arr_value_output_3.append(all_quantity)
arr_headings_output_3.append(f'F_abk')
arr_value_output_3.append(F_abk)
arr_headings_output_3.append(f'F_proiz_korp')
arr_value_output_3.append(F_proiz_korp)
arr_headings_output_3.append(f'F_x_all')
arr_value_output_3.append(F_x_all)
arr_headings_output_3.append(f'F_kpp')
arr_value_output_3.append(F_kpp)

book = openpyxl.Workbook()
data_output = book.active
for i in range(len(arr_headings_output_1)):
    data_output.cell(row=1, column=i+1).value = arr_headings_output_1[i]
    data_output.cell(row=2, column=i+1).value = arr_value_output_1[i]
book.save('output.xlsx')
book.close()

book = openpyxl.Workbook()
data_output = book.active
for i in range(len(arr_headings_output_2)):

    data_output.cell(row=1, column=i+1).value = arr_headings_output_2[i]
    data_output.cell(row=2, column=i+1).value = arr_value_output_2[i]
book.save('output2.xlsx')
book.close()


book = openpyxl.Workbook()
data_output = book.active
for i in range(len(arr_headings_output_3)):
    data_output.cell(row=1, column=i+1).value = arr_headings_output_3[i]
    data_output.cell(row=2, column=i+1).value = arr_value_output_3[i]
book.save('output3.xlsx')
book.close()
