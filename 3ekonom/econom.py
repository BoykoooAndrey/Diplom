import statistics
import car
import random
import openpyxl
import math

# vid_to_tr == 1 - TO-1; 2 - TO-2;  3 - TR

uaz_3163 = car.Vehicle("УАЗ-3163", 29, 2.7, 20, 'автомобили легковые', 'базовый автомобиль', 122.24, 300000, 4.647, 1.929)
kamaz_43502 = car.Vehicle("КАМАЗ-43502", 124, 4, 105, 'автомобили грузовые', 'базовый автомобиль', 89.85, 300000, 7.570, 2.5)
kamaz_43118 = car.Vehicle("КАМАЗ-43118", 144, 10, 93, 'автомобили грузовые', 'седельные тягачи', 107, 300000, 8.835, 2.5)
politrans = car.Vehicle("ПОЛИТРАНС-94163", 61, 40, 32, 'полуприцепы тяжеловозы', 'полуприцепы', 113.01, 250000, 15.130, 2.550)
cars = [uaz_3163, kamaz_43502, kamaz_43118, politrans]
LEN_CAR = len(cars)
L_g_do_rec=[26245.23, 19846.45, 21554.3, 25879.32]
L_g=[28054.08, 21995.28, 25647.9, 27953.02]
for i in range(LEN_CAR):
    L_g_do_rec[i] = L_g_do_rec[i] * cars[i].quantity
    L_g[i] = L_g[i] * cars[i].quantity



Lg_all_do_rec = sum(L_g_do_rec)
def find_k_3(weather):
    for i in range(18, 24):
        if weather in (table_4[i+1][0].value.split(';')):
            return float(table_4[i+1][1].value)


def find_length_proetk_zd(S_proetk_zd):
    for i in range(6, 14):
        if S_proetk_zd % i == 0:
            return i


init_data = openpyxl.open('init_data.xlsx', read_only=True)
table_1 = init_data.worksheets[0]
table_2 = init_data.worksheets[1]
table_3 = init_data.worksheets[2]
table_4 = init_data.worksheets[3]

arr_headings_output_tregub = []
arr_value_output_tregub = []

for i in range(LEN_CAR):
    arr_headings_output_tregub.append(f'cars{i}')
    arr_value_output_tregub.append(cars[i].name)

VARIANT = 10

vacation_days = 28  # Количество дней отпуска
calendar_days = 365  # Количество календарных дней
weekend_days = 118  # Количество выходных дней
weather = table_1[VARIANT + 1][32].value
if weather == 'холодные':
    vacation_days = 44
if weather == 'очень холодные':
    vacation_days = 52

# 1. РАСЧЕТ ПЛАНОВОЙ ЧИСЛЕННОСТИ РАБОТНИКОВ
# 1.2. Расчет численности младшего обслуживающего персонала
N_to1 = table_2[26][13].value
N_to2 = table_2[27][13].value
N_tr = table_2[28][13].value
N_rr = table_2[29][13].value



def append_1():
    arr_headings_output_tregub.append('N_to1')
    arr_value_output_tregub.append(N_to1)
    arr_headings_output_tregub.append('N_to2')
    arr_value_output_tregub.append(N_to2)
    arr_headings_output_tregub.append('N_tr')
    arr_value_output_tregub.append(N_tr)
    arr_headings_output_tregub.append('N_rr')
    arr_value_output_tregub.append(N_rr)
append_1()

all_quantity = 0
for i in range(LEN_CAR):
    all_quantity = all_quantity + cars[i].quantity

# 2. РАСЧЕТ ПЛАНОВОГО ФОНДА ЗАРАБОТНОЙ ПЛАТЫ
# 2.1. Определение средней тарифной ставки по видам воздействий (ТО-1, ТО-2, ТР)
tarif_3_raz = table_1[VARIANT + 1][4].value
tarif_4_raz = table_1[VARIANT + 1][5].value
tarif_5_raz = table_1[VARIANT + 1][6].value
tarif_6_raz = table_1[VARIANT + 1][7].value

workers_to_1_do_rec = []  # В массиве идет по порядку 3,4,5,6 разряд
workers_to_2_do_rec = []  # В массиве идет по порядку 3,4,5,6 разряд
workers_TR_do_rec = []  # В массиве идет по порядку 3,4,5,6 разряд

workers_to_1 = []  # В массиве идет по порядку 3,4,5,6 разряд
workers_to_2 = []  # В массиве идет по порядку 3,4,5,6 разряд
workers_TR = []  # В массиве идет по порядку 3,4,5,6 разряд
for i in range(9, 13):
    workers_to_1_do_rec.append(int(table_2[22][i].value))
    workers_to_2_do_rec.append(int(table_2[23][i].value))
    workers_TR_do_rec.append(int(table_2[24][i].value))
    
    workers_to_1.append(int(table_2[26][i].value))
    workers_to_2.append(int(table_2[27][i].value))
    workers_TR.append(int(table_2[28][i].value))

sred_tarif_to_1_do_rec = round((workers_to_1_do_rec[0] * tarif_3_raz + workers_to_1_do_rec[1] *
                         tarif_4_raz + workers_to_1_do_rec[2] * tarif_5_raz + workers_to_1_do_rec[3] * tarif_6_raz) / sum(workers_to_1_do_rec), 2)
sred_tarif_to_2_do_rec = round((workers_to_2_do_rec[0] * tarif_3_raz + workers_to_2_do_rec[1] *
                         tarif_4_raz + workers_to_2_do_rec[2] * tarif_5_raz + workers_to_2_do_rec[3] * tarif_6_raz) / sum(workers_to_2_do_rec), 2)
sred_tarif_tr_do_rec = round((workers_TR_do_rec[0] * tarif_3_raz + workers_TR_do_rec[1] *
                       tarif_4_raz + workers_TR_do_rec[2] * tarif_5_raz + workers_TR_do_rec[3] * tarif_6_raz) / sum(workers_TR_do_rec), 2)

sred_tarif_to_1 = round((workers_to_1[0] * tarif_3_raz + workers_to_1[1] *
                         tarif_4_raz + workers_to_1[2] * tarif_5_raz + workers_to_1[3] * tarif_6_raz) / sum(workers_to_1), 2)
sred_tarif_to_2 = round((workers_to_2[0] * tarif_3_raz + workers_to_2[1] *
                         tarif_4_raz + workers_to_2[2] * tarif_5_raz + workers_to_2[3] * tarif_6_raz) / sum(workers_to_2), 2)
sred_tarif_tr = round((workers_TR[0] * tarif_3_raz + workers_TR[1] *
                       tarif_4_raz + workers_TR[2] * tarif_5_raz + workers_TR[3] * tarif_6_raz) / sum(workers_TR), 2)

# 2.2 Определение заработной платы ремонтных рабочих по тарифу по видам воздействий
# 2.2.1 Фонд заработной платы ремонтных рабочих
T_to_1_do_rec = float(table_1[VARIANT + 1][19].value)
T_to_2_do_rec = float(table_1[VARIANT + 1][20].value)
T_tr_do_rec = float(table_1[VARIANT + 1][21].value)

T_to_1 = float(table_1[VARIANT + 1][2].value)
T_to_2 = float(table_1[VARIANT + 1][3].value)
T_tr = float(table_1[VARIANT + 1][1].value)



FZP_to_1_do_rec = round(sred_tarif_to_1_do_rec*T_to_1_do_rec*1.5*2.2, 2)
FZP_to_2_do_rec = round(sred_tarif_to_2_do_rec*T_to_2_do_rec*1.5*2.2, 2)
FZP_tr_do_rec = round(sred_tarif_tr_do_rec*T_tr_do_rec*1.5*2.2, 2)

FZP_to_1 = round(sred_tarif_to_1*T_to_1*1.5*2.2, 2)
FZP_to_2 = round(sred_tarif_to_2*T_to_2*1.5*2.2, 2)
FZP_tr = round(sred_tarif_tr*T_tr*1.5*2.2, 2)

def append_FZP():
    arr_headings_output_tregub.append('Lg_all_do_rec')
    arr_value_output_tregub.append(Lg_all_do_rec)
    arr_headings_output_tregub.append('all_quantity')
    arr_value_output_tregub.append(all_quantity)

    arr_headings_output_tregub.append('T_to_1_do_rec')
    arr_value_output_tregub.append(T_to_1_do_rec)
    arr_headings_output_tregub.append('T_to_2_do_rec')
    arr_value_output_tregub.append(T_to_2_do_rec)
    arr_headings_output_tregub.append('T_tr_do_rec')
    arr_value_output_tregub.append(T_tr_do_rec)
    arr_headings_output_tregub.append('T_all_do_rec')
    arr_value_output_tregub.append(T_to_1_do_rec+T_to_2_do_rec+T_tr_do_rec)
    arr_headings_output_tregub.append('FZP_to_1_do_rec')
    arr_value_output_tregub.append(FZP_to_1_do_rec)
    arr_headings_output_tregub.append('FZP_to_2_do_rec')
    arr_value_output_tregub.append(FZP_to_2_do_rec)
    arr_headings_output_tregub.append('FZP_tr_do_rec')
    arr_value_output_tregub.append(FZP_tr_do_rec)

    arr_headings_output_tregub.append('T_to_1')
    arr_value_output_tregub.append(T_to_1)
    arr_headings_output_tregub.append('T_to_2')
    arr_value_output_tregub.append(T_to_2)
    arr_headings_output_tregub.append('T_tr')
    arr_value_output_tregub.append(T_tr)
    arr_headings_output_tregub.append('T_all')
    arr_value_output_tregub.append(T_to_1+T_to_2+T_tr)
    arr_headings_output_tregub.append('FZP_to_1')
    arr_value_output_tregub.append(FZP_to_1)
    arr_headings_output_tregub.append('FZP_to_2')
    arr_value_output_tregub.append(FZP_to_2)
    arr_headings_output_tregub.append('FZP_tr')
    arr_value_output_tregub.append(FZP_tr)
append_FZP()
DZP_to1_do_rec = round(
    (vacation_days/(calendar_days - weekend_days - vacation_days))*FZP_to_1_do_rec, 2)
DZP_to2_do_rec = round(
    (vacation_days/(calendar_days - weekend_days - vacation_days))*FZP_to_2_do_rec, 2)
DZP_tr_do_rec = round(
    (vacation_days/(calendar_days - weekend_days - vacation_days))*FZP_tr_do_rec, 2)

DZP_to1 = round(
    (vacation_days/(calendar_days - weekend_days - vacation_days))*FZP_to_1, 2)
DZP_to2 = round(
    (vacation_days/(calendar_days - weekend_days - vacation_days))*FZP_to_2, 2)
DZP_tr = round(
    (vacation_days/(calendar_days - weekend_days - vacation_days))*FZP_tr, 2)

def append_FZP():
    arr_headings_output_tregub.append('calendar_days')
    arr_value_output_tregub.append(calendar_days)
    arr_headings_output_tregub.append('weekend_days')
    arr_value_output_tregub.append(weekend_days)
    arr_headings_output_tregub.append('vacation_days')
    arr_value_output_tregub.append(vacation_days)
    
    arr_headings_output_tregub.append('DZP_to1_do_rec')
    arr_value_output_tregub.append(DZP_to1_do_rec)
    arr_headings_output_tregub.append('DZP_to2_do_rec')
    arr_value_output_tregub.append(DZP_to2_do_rec)
    arr_headings_output_tregub.append('DZP_tr_do_rec')
    arr_value_output_tregub.append(DZP_tr_do_rec)
    
    arr_headings_output_tregub.append('DZP_to1')
    arr_value_output_tregub.append(DZP_to1)
    arr_headings_output_tregub.append('DZP_to2')
    arr_value_output_tregub.append(DZP_to2)
    arr_headings_output_tregub.append('DZP_tr')
    arr_value_output_tregub.append(DZP_tr)
append_FZP()

OFZP_to1_do_rec = FZP_to_1_do_rec + DZP_to1_do_rec
OFZP_to2_do_rec =FZP_to_2_do_rec + DZP_to2_do_rec
OFZP_tr_do_rec = FZP_tr_do_rec + DZP_tr_do_rec

OFZP_to1 = FZP_to_1 + DZP_to1
OFZP_to2 =FZP_to_2 + DZP_to2
OFZP_tr = FZP_tr + DZP_tr

# 2.9.6. Отчисление в соц. фонды по видам воздействия
SOC_OTC_to1_do_rec = round((OFZP_to1_do_rec * 30) / 100, 2)
SOC_OTC_to2_do_rec = round((OFZP_to2_do_rec * 30) / 100, 2)
SOC_OTC_tr_do_rec = round((OFZP_tr_do_rec * 30) / 100, 2)

SOC_OTC_to1 = round((OFZP_to1 * 30) / 100, 2)
SOC_OTC_to2 = round((OFZP_to2 * 30) / 100, 2)
SOC_OTC_tr = round((OFZP_tr * 30) / 100, 2)
# 2.9.7. Фонд заработной платы с отчислениями по видам воздействия
OFZP_all_with_otc_to1_do_rec = OFZP_to1_do_rec + SOC_OTC_to1_do_rec
OFZP_all_with_otc_to2_do_rec = OFZP_to2_do_rec + SOC_OTC_to2_do_rec
OFZP_all_with_otc_tr_do_rec = OFZP_tr_do_rec + SOC_OTC_tr_do_rec
OFZP_all_with_otc_all_do_rec = OFZP_all_with_otc_to1_do_rec + OFZP_all_with_otc_to2_do_rec + OFZP_all_with_otc_tr_do_rec

OFZP_all_with_otc_to1 = OFZP_to1 + SOC_OTC_to1
OFZP_all_with_otc_to2 = OFZP_to2 + SOC_OTC_to2
OFZP_all_with_otc_tr = OFZP_tr + SOC_OTC_tr
OFZP_all_with_otc_all = OFZP_all_with_otc_to1 + OFZP_all_with_otc_to2 + OFZP_all_with_otc_tr

def append_OFZP():
    arr_headings_output_tregub.append('OFZP_to1_do_rec')
    arr_value_output_tregub.append(OFZP_to1_do_rec)
    arr_headings_output_tregub.append('OFZP_to2_do_rec')
    arr_value_output_tregub.append(OFZP_to2_do_rec)
    arr_headings_output_tregub.append('OFZP_tr_do_rec')
    arr_value_output_tregub.append(OFZP_tr_do_rec)
    arr_headings_output_tregub.append('SOC_OTC_to1_do_rec')
    arr_value_output_tregub.append(SOC_OTC_to1_do_rec)
    arr_headings_output_tregub.append('SOC_OTC_to2_do_rec')
    arr_value_output_tregub.append(SOC_OTC_to2_do_rec)
    arr_headings_output_tregub.append('SOC_OTC_tr_do_rec')
    arr_value_output_tregub.append(SOC_OTC_tr_do_rec)
    arr_headings_output_tregub.append('OFZP_all_with_otc_to1_do_rec')
    arr_value_output_tregub.append(OFZP_all_with_otc_to1_do_rec)
    arr_headings_output_tregub.append('OFZP_all_with_otc_to2_do_rec')
    arr_value_output_tregub.append(OFZP_all_with_otc_to2_do_rec)
    arr_headings_output_tregub.append('OFZP_all_with_otc_tr_do_rec')
    arr_value_output_tregub.append(OFZP_all_with_otc_tr_do_rec)
    arr_headings_output_tregub.append('OFZP_all_with_otc_all_do_rec')
    arr_value_output_tregub.append(OFZP_all_with_otc_all_do_rec)

    arr_headings_output_tregub.append('OFZP_to1')
    arr_value_output_tregub.append(OFZP_to1)
    arr_headings_output_tregub.append('OFZP_to2')
    arr_value_output_tregub.append(OFZP_to2)
    arr_headings_output_tregub.append('OFZP_tr')
    arr_value_output_tregub.append(OFZP_tr)
    arr_headings_output_tregub.append('SOC_OTC_to1')
    arr_value_output_tregub.append(SOC_OTC_to1)
    arr_headings_output_tregub.append('SOC_OTC_to2')
    arr_value_output_tregub.append(SOC_OTC_to2)
    arr_headings_output_tregub.append('SOC_OTC_tr')
    arr_value_output_tregub.append(SOC_OTC_tr)
    arr_headings_output_tregub.append('OFZP_all_with_otc_to1')
    arr_value_output_tregub.append(OFZP_all_with_otc_to1)
    arr_headings_output_tregub.append('OFZP_all_with_otc_to2')
    arr_value_output_tregub.append(OFZP_all_with_otc_to2)
    arr_headings_output_tregub.append('OFZP_all_with_otc_tr')
    arr_value_output_tregub.append(OFZP_all_with_otc_tr)
    arr_headings_output_tregub.append('OFZP_all_with_otc_all')
    arr_value_output_tregub.append(OFZP_all_with_otc_all)
append_OFZP()

def append_2_1():

    arr_headings_output_tregub.append('tarif_3_raz')
    arr_value_output_tregub.append(tarif_3_raz)
    arr_headings_output_tregub.append('tarif_4_raz')
    arr_value_output_tregub.append(tarif_4_raz)
    arr_headings_output_tregub.append('tarif_5_raz')
    arr_value_output_tregub.append(tarif_5_raz)
    arr_headings_output_tregub.append('tarif_6_raz')
    arr_value_output_tregub.append(tarif_6_raz)

    arr_headings_output_tregub.append('workers_to_1_3_raz_do_rec')
    arr_value_output_tregub.append(workers_to_1_do_rec[0])
    arr_headings_output_tregub.append('workers_to_1_4_raz_do_rec')
    arr_value_output_tregub.append(workers_to_1_do_rec[1])
    arr_headings_output_tregub.append('workers_to_1_5_raz_do_rec')
    arr_value_output_tregub.append(workers_to_1_do_rec[2])
    arr_headings_output_tregub.append('workers_to_1_6_raz_do_rec')
    arr_value_output_tregub.append(workers_to_1_do_rec[3])
    arr_headings_output_tregub.append('sred_tarif_to_1_do_rec')
    arr_value_output_tregub.append(sred_tarif_to_1_do_rec)
    arr_headings_output_tregub.append('workers_to_2_3_raz_do_rec')
    arr_value_output_tregub.append(workers_to_2_do_rec[0])
    arr_headings_output_tregub.append('workers_to_2_4_raz_do_rec')
    arr_value_output_tregub.append(workers_to_2_do_rec[1])
    arr_headings_output_tregub.append('workers_to_2_5_raz_do_rec')
    arr_value_output_tregub.append(workers_to_2_do_rec[2])
    arr_headings_output_tregub.append('workers_to_2_6_raz_do_rec')
    arr_value_output_tregub.append(workers_to_2_do_rec[3])
    arr_headings_output_tregub.append('sred_tarif_to_2_do_rec')
    arr_value_output_tregub.append(sred_tarif_to_2_do_rec)
    arr_headings_output_tregub.append('workers_TR_3_raz_do_rec')
    arr_value_output_tregub.append(workers_TR_do_rec[0])
    arr_headings_output_tregub.append('workers_TR_4_raz_do_rec')
    arr_value_output_tregub.append(workers_TR_do_rec[1])
    arr_headings_output_tregub.append('workers_TR_5_raz_do_rec')
    arr_value_output_tregub.append(workers_TR_do_rec[2])
    arr_headings_output_tregub.append('workers_TR_6_raz_do_rec')
    arr_value_output_tregub.append(workers_TR_do_rec[3])
    arr_headings_output_tregub.append('sred_tarif_TR_do_rec')
    arr_value_output_tregub.append(sred_tarif_tr_do_rec)

    arr_headings_output_tregub.append('workers_to_1_3_raz')
    arr_value_output_tregub.append(workers_to_1[0])
    arr_headings_output_tregub.append('workers_to_1_4_raz')
    arr_value_output_tregub.append(workers_to_1[1])
    arr_headings_output_tregub.append('workers_to_1_5_raz')
    arr_value_output_tregub.append(workers_to_1[2])
    arr_headings_output_tregub.append('workers_to_1_6_raz')
    arr_value_output_tregub.append(workers_to_1[3])
    arr_headings_output_tregub.append('sred_tarif_to_1')
    arr_value_output_tregub.append(sred_tarif_to_1)
    arr_headings_output_tregub.append('workers_to_2_3_raz')
    arr_value_output_tregub.append(workers_to_2[0])
    arr_headings_output_tregub.append('workers_to_2_4_raz')
    arr_value_output_tregub.append(workers_to_2[1])
    arr_headings_output_tregub.append('workers_to_2_5_raz')
    arr_value_output_tregub.append(workers_to_2[2])
    arr_headings_output_tregub.append('workers_to_2_6_raz')
    arr_value_output_tregub.append(workers_to_2[3])
    arr_headings_output_tregub.append('sred_tarif_to_2')
    arr_value_output_tregub.append(sred_tarif_to_2)
    arr_headings_output_tregub.append('workers_TR_3_raz')
    arr_value_output_tregub.append(workers_TR[0])
    arr_headings_output_tregub.append('workers_TR_4_raz')
    arr_value_output_tregub.append(workers_TR[1])
    arr_headings_output_tregub.append('workers_TR_5_raz')
    arr_value_output_tregub.append(workers_TR[2])
    arr_headings_output_tregub.append('workers_TR_6_raz')
    arr_value_output_tregub.append(workers_TR[3])
    arr_headings_output_tregub.append('sred_tarif_TR')
    arr_value_output_tregub.append(sred_tarif_tr)
append_2_1()


Z_naklad_to_1_do_rec = round(OFZP_all_with_otc_to1_do_rec * 0.5, 2)
Z_naklad_to_2_do_rec = round(OFZP_all_with_otc_to2_do_rec * 0.5, 2)
Z_naklad_tr_do_rec = round(OFZP_all_with_otc_tr_do_rec * 0.5, 2)

Z_naklad_to_1 = round(OFZP_all_with_otc_to1 * 0.5, 2)
Z_naklad_to_2 = round(OFZP_all_with_otc_to2 * 0.5, 2)
Z_naklad_tr = round(OFZP_all_with_otc_tr * 0.5, 2)
# 3.6. Затраты на расходные материалы и запасные части для ремонтной зоны
def find_H_m(TYPE_VEHICLE, vid_to_tr):
    if TYPE_VEHICLE == 'автомобили легковые':
        return table_3[vid_to_tr+1][1].value
    elif TYPE_VEHICLE == 'автомобили грузовые':
        return table_3[vid_to_tr+1][2].value
    elif TYPE_VEHICLE == 'автобусы':
        return table_3[vid_to_tr+1][3].value
    else:
        return table_3[vid_to_tr+1][2].value
def find_H_zch(TYPE_VEHICLE):
    if TYPE_VEHICLE == 'автомобили легковые':
        return table_3[7][1].value
    elif TYPE_VEHICLE == 'автомобили грузовые':
        return table_3[7][2].value
    elif TYPE_VEHICLE == 'автобусы':
        return table_3[7][3].value
    else:
        return table_3[7][2].value

L_g_all_do_rec = sum(L_g_do_rec)
Lg_all = sum(L_g)
category = table_1[VARIANT + 1][31].value
k_1 = table_4[category+1][1].value
k_3 = find_k_3(weather)

Z_mat_to_1_all_do_rec = []
Z_mat_to_2_all_do_rec = []
Z_mat_tr_all_do_rec = []
Z_zch_all_do_rec = []


Z_mat_to_1_all = []
Z_mat_to_2_all = []
Z_mat_tr_all = []
Z_zch_all = []

for i in range(LEN_CAR):
    if cars[i].modif_vihicle == 'седельные тягачи':
        k_2 = 1.05
    else:
        k_2 = 1
    
    norma_z_mat_to1 = find_H_m(cars[i].type_vihicle, 1)
    norma_z_mat_to2 = find_H_m(cars[i].type_vihicle, 2)

    Z_mat_to_1_do_rec = round((norma_z_mat_to1 * L_g_do_rec[i])/1000, 2)
    Z_mat_to_2_do_rec = round((norma_z_mat_to2 * L_g_do_rec[i])/1000, 2)
    Z_mat_to_1_all_do_rec.append(Z_mat_to_1_do_rec)
    Z_mat_to_2_all_do_rec.append(Z_mat_to_2_do_rec)

    Z_mat_to_1 = round((norma_z_mat_to1 * L_g[i])/1000, 2)
    Z_mat_to_2 = round((norma_z_mat_to2 * L_g[i])/1000, 2)
    Z_mat_to_1_all.append(Z_mat_to_1)
    Z_mat_to_2_all.append(Z_mat_to_2)

    norma_z_mat_tr = find_H_m(cars[i].type_vihicle, 3)

    Z_mat_tr_do_rec = round((norma_z_mat_tr * L_g_do_rec[i])/1000, 2)
    Z_mat_tr_all_do_rec.append(Z_mat_tr_do_rec)

    Z_mat_tr = round((norma_z_mat_tr * L_g[i])/1000, 2)
    Z_mat_tr_all.append(Z_mat_tr)

    norma_z_zch = find_H_zch(cars[i].type_vihicle)

    Z_zch_do_rec = round(((norma_z_zch * L_g_do_rec[i])/1000)*k_1*k_2*k_3, 2)
    Z_zch_all_do_rec.append(Z_zch_do_rec)

    Z_zch = round(((norma_z_zch * L_g[i])/1000)*k_1*k_2*k_3, 2)
    Z_zch_all.append(Z_zch)

    arr_headings_output_tregub.append(f'norma_z_mat_to1{i}')
    arr_value_output_tregub.append(norma_z_mat_to1)
    arr_headings_output_tregub.append(f'norma_z_mat_to2{i}')
    arr_value_output_tregub.append(norma_z_mat_to2)

    arr_headings_output_tregub.append(f'Z_mat_to_1_do_rec{i}')
    arr_value_output_tregub.append(Z_mat_to_1_do_rec)
    arr_headings_output_tregub.append(f'Z_mat_to_2_do_rec{i}')
    arr_value_output_tregub.append(Z_mat_to_2_do_rec)

    arr_headings_output_tregub.append(f'Z_mat_to_1{i}')
    arr_value_output_tregub.append(Z_mat_to_1)
    arr_headings_output_tregub.append(f'Z_mat_to_2{i}')
    arr_value_output_tregub.append(Z_mat_to_2)
    
    arr_headings_output_tregub.append(f'L_g_do_rec{i}')
    arr_value_output_tregub.append(L_g_do_rec[i])
    arr_headings_output_tregub.append(f'L_g{i}')
    arr_value_output_tregub.append(L_g[i])

    arr_headings_output_tregub.append(f'norma_z_mat_tr{i}')
    arr_value_output_tregub.append(norma_z_mat_tr)
    arr_headings_output_tregub.append(f'Z_mat_tr_do_rec{i}')
    arr_value_output_tregub.append(Z_mat_tr_do_rec)
    arr_headings_output_tregub.append(f'Z_mat_tr{i}')
    arr_value_output_tregub.append(Z_mat_tr)

    arr_headings_output_tregub.append(f'k_2{i}')
    arr_value_output_tregub.append(k_2)
    arr_headings_output_tregub.append(f'norma_z_zch{i}')
    arr_value_output_tregub.append(norma_z_zch)
    arr_headings_output_tregub.append(f'Z_zch_do_rec{i}')
    arr_value_output_tregub.append(Z_zch_do_rec)
    arr_headings_output_tregub.append(f'Z_zch{i}')
    arr_value_output_tregub.append(Z_zch)

def append_3():
    arr_headings_output_tregub.append('Z_naklad_to_1_do_rec')
    arr_value_output_tregub.append(Z_naklad_to_1_do_rec)
    arr_headings_output_tregub.append('Z_naklad_to_2_do_rec')
    arr_value_output_tregub.append(Z_naklad_to_2_do_rec)
    arr_headings_output_tregub.append('Z_naklad_tr_do_rec')
    arr_value_output_tregub.append(Z_naklad_tr_do_rec)
    
    arr_headings_output_tregub.append('Z_naklad_to_1')
    arr_value_output_tregub.append(Z_naklad_to_1)
    arr_headings_output_tregub.append('Z_naklad_to_2')
    arr_value_output_tregub.append(Z_naklad_to_2)
    arr_headings_output_tregub.append('Z_naklad_tr')
    arr_value_output_tregub.append(Z_naklad_tr)

    arr_headings_output_tregub.append('k_1')
    arr_value_output_tregub.append(k_1)
    arr_headings_output_tregub.append('k_3')
    arr_value_output_tregub.append(k_3)

    arr_headings_output_tregub.append('Z_mat_tr_all')
    arr_value_output_tregub.append(sum(Z_mat_tr_all))

    arr_headings_output_tregub.append('Z_zch_all')
    arr_value_output_tregub.append(sum(Z_zch_all))


append_3()

# 7.4. Затраты на приобретение оборудования
zatrati_na_new_oborud = round(table_1[VARIANT + 1][45].value, 2)
Z_mont_dem_trans = round(zatrati_na_new_oborud * 0.25, 2)
zatrati_na_new_oborud_with_mont = zatrati_na_new_oborud + Z_mont_dem_trans
def append_7():
    arr_headings_output_tregub.append('zatrati_na_new_oborud')
    arr_value_output_tregub.append(zatrati_na_new_oborud)
    arr_headings_output_tregub.append('Z_mont_dem_trans')
    arr_value_output_tregub.append(Z_mont_dem_trans)
    arr_headings_output_tregub.append('zatrati_na_new_oborud_with_mont')
    arr_value_output_tregub.append(zatrati_na_new_oborud_with_mont)
append_7()



# 6 СОСТАВЛЕНИЕ СМЕТЫ ЗАТРАТ И КАЛЬКУЛЯЦИИ СЕБЕСТОИМОСТИ ТО И ТР
N_to_1_do_rec = round(table_1[VARIANT + 1][40].value)
S_to_1_on_1000km_do_rec = round(((OFZP_all_with_otc_to1_do_rec + sum(Z_mat_to_1_all_do_rec) + Z_naklad_to_1_do_rec) / Lg_all_do_rec) * 1000, 2)
S_1_to_1_do_rec = round(((OFZP_all_with_otc_to1_do_rec + sum(Z_mat_to_1_all_do_rec) + Z_naklad_to_1_do_rec) / N_to_1_do_rec), 2)
S_to_1_on_1chel_ch_do_rec = round(((OFZP_all_with_otc_to1_do_rec + sum(Z_mat_to_1_all_do_rec) + Z_naklad_to_1_do_rec) / T_to_1_do_rec), 2)

N_to_1 = round(table_1[VARIANT + 1][42].value)
S_to_1_on_1000km = round(((OFZP_all_with_otc_to1 + sum(Z_mat_to_1_all) + Z_naklad_to_1) / Lg_all) * 1000, 2)
S_1_to_1 = round(((OFZP_all_with_otc_to1 + sum(Z_mat_to_1_all) + Z_naklad_to_1) / N_to_1), 2)
S_to_1_on_1chel_ch = round(((OFZP_all_with_otc_to1 + sum(Z_mat_to_1_all) + Z_naklad_to_1) / T_to_1), 2)

N_to_2_do_rec = round(table_1[VARIANT + 1][41].value)
S_to_2_on_1000km_do_rec = round(((OFZP_all_with_otc_to2_do_rec + sum(Z_mat_to_2_all_do_rec) + Z_naklad_to_2_do_rec) / Lg_all_do_rec) * 1000, 2)
S_1_to_2_do_rec = round(((OFZP_all_with_otc_to2_do_rec + sum(Z_mat_to_2_all_do_rec) + Z_naklad_to_2_do_rec) / N_to_2_do_rec), 2)
S_to_2_on_1chel_ch_do_rec = round(((OFZP_all_with_otc_to2_do_rec + sum(Z_mat_to_2_all_do_rec) + Z_naklad_to_2_do_rec) / T_to_2_do_rec), 2)

N_to_2 = round(table_1[VARIANT + 1][43].value)
S_to_2_on_1000km = round(((OFZP_all_with_otc_to2 + sum(Z_mat_to_2_all) + Z_naklad_to_2) / Lg_all) * 1000, 2)
S_1_to_2 = round(((OFZP_all_with_otc_to2 + sum(Z_mat_to_2_all) + Z_naklad_to_2) / N_to_2), 2)
S_to_2_on_1chel_ch = round(((OFZP_all_with_otc_to2 + sum(Z_mat_to_2_all) + Z_naklad_to_2) / T_to_2), 2)

S_1_chel_tr_do_rec = round((OFZP_all_with_otc_tr_do_rec + sum(Z_mat_tr_all_do_rec) + sum(Z_zch_all_do_rec) + Z_naklad_tr_do_rec) / T_tr_do_rec, 2)
S_tr_on_1000km_do_rec = round(((OFZP_all_with_otc_tr_do_rec + sum(Z_mat_tr_all_do_rec) + sum(Z_zch_all_do_rec) + Z_naklad_tr_do_rec) / Lg_all_do_rec) * 1000, 2)

S_1_chel_tr = round((OFZP_all_with_otc_tr + sum(Z_mat_tr_all) + sum(Z_zch_all) + Z_naklad_tr) / T_tr, 2)
S_tr_on_1000km = round(((OFZP_all_with_otc_tr + sum(Z_mat_tr_all) + sum(Z_zch_all) + Z_naklad_tr) / Lg_all) * 1000, 2)


#3.1 Снижение себестоимости То и Р 
P_cc_to_1 = round((S_1_to_1_do_rec/S_1_to_1 - 1) * 100, 2)
P_cc_to_2 = round((S_1_to_2_do_rec/S_1_to_2 - 1) * 100, 2)
delta_C_to_1 = (S_1_to_1_do_rec - S_1_to_1) * N_to_1
delta_C_to_2 = (S_1_to_2_do_rec - S_1_to_2) * N_to_2
delta_C_tr = round((S_tr_on_1000km_do_rec - S_tr_on_1000km) * (Lg_all/1000), 2)
delta_C_to_tr = delta_C_to_1 + delta_C_to_2 + delta_C_tr
T_okup = round(zatrati_na_new_oborud/ delta_C_to_tr, 2)


def append_6():
    arr_headings_output_tregub.append('S_1_to_1_do_rec')
    arr_value_output_tregub.append(S_1_to_1_do_rec)
    arr_headings_output_tregub.append('S_1_to_1')
    arr_value_output_tregub.append(S_1_to_1)
    arr_headings_output_tregub.append('P_cc_to_1')
    arr_value_output_tregub.append(P_cc_to_1)
    arr_headings_output_tregub.append('S_1_to_2_do_rec')
    arr_value_output_tregub.append(S_1_to_2_do_rec)
    arr_headings_output_tregub.append('S_1_to_2')
    arr_value_output_tregub.append(S_1_to_2)
    arr_headings_output_tregub.append('P_cc_to_2')
    arr_value_output_tregub.append(P_cc_to_2)
    arr_headings_output_tregub.append('N_to_1')
    arr_value_output_tregub.append(N_to_1)
    arr_headings_output_tregub.append('delta_C_to_1')
    arr_value_output_tregub.append(delta_C_to_1)
    arr_headings_output_tregub.append('N_to_2')
    arr_value_output_tregub.append(N_to_2)
    arr_headings_output_tregub.append('delta_C_to_2')
    arr_value_output_tregub.append(delta_C_to_2)
    arr_headings_output_tregub.append('S_tr_on_1000km_do_rec')
    arr_value_output_tregub.append(S_tr_on_1000km_do_rec)
    arr_headings_output_tregub.append('S_tr_on_1000km')
    arr_value_output_tregub.append(S_tr_on_1000km)
    arr_headings_output_tregub.append('delta_C_tr')
    arr_value_output_tregub.append(delta_C_tr)
    arr_headings_output_tregub.append('delta_C_to_tr')
    arr_value_output_tregub.append(delta_C_to_tr)

    arr_headings_output_tregub.append('T_okup')
    arr_value_output_tregub.append(T_okup)
    arr_headings_output_tregub.append('Lg_all')
    arr_value_output_tregub.append(Lg_all)
    arr_headings_output_tregub.append('zatrati_na_new_oborud_with_mont')
    arr_value_output_tregub.append(zatrati_na_new_oborud_with_mont)
append_6()


def append_for_ishod_date():
    arr_headings_output_tregub.append('T_sm')
    arr_value_output_tregub.append(table_1[VARIANT + 1][8].value)
    arr_headings_output_tregub.append('A_m_day_job')
    arr_value_output_tregub.append(table_1[VARIANT + 1][10].value)
    arr_headings_output_tregub.append('S_ofis_pom')
    arr_value_output_tregub.append(table_1[VARIANT + 1][12].value)
    arr_headings_output_tregub.append('Kol-vo postov')
    arr_value_output_tregub.append(table_1[VARIANT + 1][25].value)
    arr_headings_output_tregub.append('day_km')
    arr_value_output_tregub.append(table_1[VARIANT + 1][29].value)
    arr_headings_output_tregub.append('a_ttt')
    arr_value_output_tregub.append(table_1[VARIANT + 1][30].value)
    arr_headings_output_tregub.append('category')
    arr_value_output_tregub.append(category)
    arr_headings_output_tregub.append('weather')
    arr_value_output_tregub.append(weather)
    arr_headings_output_tregub.append('Kol-vo a_m')
    arr_value_output_tregub.append(((table_1[VARIANT + 1][33].value).split(";"))[0])
append_for_ishod_date()

def table_6_0(tables):
    tables.cell(row=4, column=1).value = OFZP_to1
    tables.cell(row=5, column=1).value = SOC_OTC_to1
    tables.cell(row=6, column=1).value = sum(Z_mat_to_1_all)
    tables.cell(row=7, column=1).value = Z_naklad_to_1
    tables.cell(row=8, column=1).value = OFZP_to1 +  SOC_OTC_to1 +  sum(Z_mat_to_1_all) + Z_naklad_to_1

    tables.cell(row=4, column=2).value = round((OFZP_to1/ Lg_all) * 1000, 2)
    tables.cell(row=5, column=2).value = round((SOC_OTC_to1  / Lg_all) * 1000, 2)
    tables.cell(row=6, column=2).value = round(((sum(Z_mat_to_1_all)) / Lg_all) * 1000, 2)
    tables.cell(row=7, column=2).value = round((Z_naklad_to_1 / Lg_all) * 1000, 2)
    tables.cell(row=8, column=2).value = round(( (OFZP_to1 +  SOC_OTC_to1 +  sum(Z_mat_to_1_all) + Z_naklad_to_1) / Lg_all) * 1000, 2)

    tables.cell(row=4, column=3).value = round((OFZP_to1 / N_to_1), 2)
    tables.cell(row=5, column=3).value = round((SOC_OTC_to1 / N_to_1), 2)
    tables.cell(row=6, column=3).value = round((sum(Z_mat_to_1_all) / N_to_1), 2)
    tables.cell(row=7, column=3).value = round((Z_naklad_to_1 / N_to_1), 2)
    tables.cell(row=8, column=3).value = round((OFZP_to1 +  SOC_OTC_to1 +  sum(Z_mat_to_1_all) + Z_naklad_to_1) / N_to_1, 2)

    tables.cell(row=4, column=4).value = round((OFZP_to1 / T_to_1), 2)
    tables.cell(row=5, column=4).value = round((SOC_OTC_to1 / T_to_1), 2)
    tables.cell(row=6, column=4).value = round((sum(Z_mat_to_1_all) / T_to_1), 2)
    tables.cell(row=7, column=4).value = round((Z_naklad_to_1 / T_to_1), 2)
    tables.cell(row=8, column=4).value = round((OFZP_to1 +  SOC_OTC_to1 +  sum(Z_mat_to_1_all) + Z_naklad_to_1) / T_to_1, 2)

    tables.cell(row=9, column=1).value = "table 6_0_0"

    tables.cell(row=4, column=5).value = OFZP_to1_do_rec
    tables.cell(row=5, column=5).value = SOC_OTC_to1_do_rec
    tables.cell(row=6, column=5).value = sum(Z_mat_to_1_all_do_rec)
    tables.cell(row=7, column=5).value = Z_naklad_to_1_do_rec
    tables.cell(row=8, column=5).value = OFZP_to1_do_rec +  SOC_OTC_to1_do_rec +  sum(Z_mat_to_1_all_do_rec) + Z_naklad_to_1_do_rec

    tables.cell(row=4, column=6).value = round((OFZP_to1_do_rec/ Lg_all_do_rec) * 1000, 2)
    tables.cell(row=5, column=6).value = round((SOC_OTC_to1_do_rec  / Lg_all_do_rec) * 1000, 2)
    tables.cell(row=6, column=6).value = round(((sum(Z_mat_to_1_all_do_rec)) / Lg_all_do_rec) * 1000, 2)
    tables.cell(row=7, column=6).value = round((Z_naklad_to_1_do_rec / Lg_all_do_rec) * 1000, 2)
    tables.cell(row=8, column=6).value = round(( (OFZP_to1_do_rec +  SOC_OTC_to1_do_rec +  sum(Z_mat_to_1_all_do_rec) + Z_naklad_to_1_do_rec) / Lg_all_do_rec) * 1000, 2)

    tables.cell(row=4, column=7).value = round((OFZP_to1_do_rec / N_to_1_do_rec), 2)
    tables.cell(row=5, column=7).value = round((SOC_OTC_to1_do_rec / N_to_1_do_rec), 2)
    tables.cell(row=6, column=7).value = round((sum(Z_mat_to_1_all_do_rec) / N_to_1_do_rec), 2)
    tables.cell(row=7, column=7).value = round((Z_naklad_to_1_do_rec / N_to_1_do_rec), 2)
    tables.cell(row=8, column=7).value = round((OFZP_to1_do_rec +  SOC_OTC_to1_do_rec +  sum(Z_mat_to_1_all_do_rec) + Z_naklad_to_1_do_rec) / N_to_1_do_rec, 2)

    tables.cell(row=4, column=8).value = round((OFZP_to1_do_rec / T_to_1_do_rec), 2)
    tables.cell(row=5, column=8).value = round((SOC_OTC_to1_do_rec / T_to_1_do_rec), 2)
    tables.cell(row=6, column=8).value = round((sum(Z_mat_to_1_all_do_rec) / T_to_1_do_rec), 2)
    tables.cell(row=7, column=8).value = round((Z_naklad_to_1_do_rec / T_to_1_do_rec), 2)
    tables.cell(row=8, column=8).value = round((OFZP_to1_do_rec +  SOC_OTC_to1_do_rec +  sum(Z_mat_to_1_all_do_rec) + Z_naklad_to_1_do_rec) / T_to_1_do_rec, 2)

    tables.cell(row=9, column=5).value = "table 6_0_1"

def table_6_1(tables):
    tables.cell(row=11, column=1).value = OFZP_to2
    tables.cell(row=12, column=1).value = SOC_OTC_to2
    tables.cell(row=13, column=1).value = sum(Z_mat_to_2_all)
    tables.cell(row=14, column=1).value = Z_naklad_to_2
    tables.cell(row=15, column=1).value = OFZP_to2 + SOC_OTC_to2 + sum(Z_mat_to_2_all) + Z_naklad_to_2

    tables.cell(row=11, column=2).value = round((OFZP_to2 / Lg_all) * 1000, 2)
    tables.cell(row=12, column=2).value = round((SOC_OTC_to2 / Lg_all) * 1000, 2)
    tables.cell(row=13, column=2).value = round(((sum(Z_mat_to_2_all)) / Lg_all) * 1000, 2)
    tables.cell(row=14, column=2).value = round(((Z_naklad_to_2) / Lg_all) * 1000, 2)
    tables.cell(row=15, column=2).value = round(((OFZP_to2 + SOC_OTC_to2 + sum(Z_mat_to_2_all) + Z_naklad_to_2) / Lg_all) * 1000, 2)

    tables.cell(row=11, column=3).value = round((OFZP_to2 / N_to_2), 2)
    tables.cell(row=12, column=3).value = round((SOC_OTC_to2 / N_to_2), 2)
    tables.cell(row=13, column=3).value = round(((sum(Z_mat_to_2_all)) / N_to_2), 2)
    tables.cell(row=14, column=3).value = round(((Z_naklad_to_2) / N_to_2), 2)
    tables.cell(row=15, column=3).value = round((OFZP_to2 + SOC_OTC_to2 + sum(Z_mat_to_2_all) + Z_naklad_to_2) / N_to_2, 2)

    tables.cell(row=11, column=4).value = round((OFZP_to2 / T_to_2), 2)
    tables.cell(row=12, column=4).value = round((SOC_OTC_to2 / T_to_2), 2)
    tables.cell(row=13, column=4).value = round(((sum(Z_mat_to_2_all)) / T_to_2), 2)
    tables.cell(row=14, column=4).value = round(((Z_naklad_to_2) / T_to_2), 2)
    tables.cell(row=15, column=4).value = round((OFZP_to2 + SOC_OTC_to2 + sum(Z_mat_to_2_all) + Z_naklad_to_2) / T_to_2, 2)

    tables.cell(row=16, column=5).value = "table 6_1_0"

    tables.cell(row=11, column=5).value = OFZP_to2_do_rec
    tables.cell(row=12, column=5).value = SOC_OTC_to2_do_rec
    tables.cell(row=13, column=5).value = sum(Z_mat_to_2_all_do_rec)
    tables.cell(row=14, column=5).value = Z_naklad_to_2_do_rec
    tables.cell(row=15, column=5).value = OFZP_to2_do_rec + SOC_OTC_to2_do_rec + sum(Z_mat_to_2_all_do_rec) + Z_naklad_to_2_do_rec

    tables.cell(row=11, column=6).value = round((OFZP_to2_do_rec / Lg_all_do_rec) * 1000, 2)
    tables.cell(row=12, column=6).value = round((SOC_OTC_to2_do_rec / Lg_all_do_rec) * 1000, 2)
    tables.cell(row=13, column=6).value = round(((sum(Z_mat_to_2_all_do_rec)) / Lg_all_do_rec) * 1000, 2)
    tables.cell(row=14, column=6).value = round(((Z_naklad_to_2_do_rec) / Lg_all_do_rec) * 1000, 2)
    tables.cell(row=15, column=6).value = round(((OFZP_to2_do_rec + SOC_OTC_to2_do_rec + sum(Z_mat_to_2_all_do_rec) + Z_naklad_to_2_do_rec) / Lg_all_do_rec) * 1000, 2)

    tables.cell(row=11, column=7).value = round((OFZP_to2_do_rec / N_to_2_do_rec), 2)
    tables.cell(row=12, column=7).value = round((SOC_OTC_to2_do_rec / N_to_2_do_rec), 2)
    tables.cell(row=13, column=7).value = round(((sum(Z_mat_to_2_all_do_rec)) / N_to_2_do_rec), 2)
    tables.cell(row=14, column=7).value = round(((Z_naklad_to_2_do_rec) / N_to_2_do_rec), 2)
    tables.cell(row=15, column=7).value = round((OFZP_to2_do_rec + SOC_OTC_to2_do_rec + sum(Z_mat_to_2_all_do_rec) + Z_naklad_to_2_do_rec) / N_to_2_do_rec, 2)

    tables.cell(row=11, column=8).value = round((OFZP_to2_do_rec / T_to_2_do_rec), 2)
    tables.cell(row=12, column=8).value = round((SOC_OTC_to2_do_rec / T_to_2_do_rec), 2)
    tables.cell(row=13, column=8).value = round(((sum(Z_mat_to_2_all_do_rec)) / T_to_2_do_rec), 2)
    tables.cell(row=14, column=8).value = round(((Z_naklad_to_2_do_rec) / T_to_2_do_rec), 2)
    tables.cell(row=15, column=8).value = round((OFZP_to2_do_rec + SOC_OTC_to2_do_rec + sum(Z_mat_to_2_all_do_rec) + Z_naklad_to_2_do_rec) / T_to_2_do_rec, 2)

    tables.cell(row=16, column=5).value = "table 6_1_1"

def table_6_2(tables):
    tables.cell(row=17, column=1).value = OFZP_tr
    tables.cell(row=18, column=1).value = SOC_OTC_tr
    tables.cell(row=19, column=1).value = sum(Z_mat_tr_all)
    tables.cell(row=20, column=1).value = sum(Z_zch_all)
    tables.cell(row=21, column=1).value = Z_naklad_tr
    tables.cell(row=22, column=1).value = OFZP_tr + SOC_OTC_tr + sum(Z_mat_tr_all) +  sum(Z_zch_all) + Z_naklad_tr

    tables.cell(row=17, column=2).value = round((OFZP_tr / Lg_all) * 1000, 2)
    tables.cell(row=18, column=2).value = round((SOC_OTC_tr / Lg_all) * 1000, 2)
    tables.cell(row=19, column=2).value = round((sum(Z_mat_tr_all) / Lg_all) * 1000, 2)
    tables.cell(row=20, column=2).value = round((sum(Z_zch_all) / Lg_all) * 1000, 2)
    tables.cell(row=21, column=2).value = round((Z_naklad_tr  / Lg_all) * 1000, 2)
    tables.cell(row=22, column=2).value = round(((OFZP_tr + SOC_OTC_tr + sum(Z_mat_tr_all) + sum(Z_zch_all) + Z_naklad_tr) / Lg_all) * 1000, 2)

    tables.cell(row=17, column=3).value = round((OFZP_tr / T_tr), 2)
    tables.cell(row=18, column=3).value = round((SOC_OTC_tr / T_tr), 2)
    tables.cell(row=19, column=3).value = round((sum(Z_mat_tr_all) / T_tr), 2)
    tables.cell(row=20, column=3).value = round((sum(Z_zch_all) / T_tr), 2)
    tables.cell(row=21, column=3).value = round((Z_naklad_tr / T_tr), 2)
    tables.cell(row=22, column=3).value = round((OFZP_tr + SOC_OTC_tr + sum(Z_mat_tr_all) + sum(Z_zch_all) + Z_naklad_tr) / T_tr, 2)

    tables.cell(row=23, column=1).value = "table 6_2_0"

    tables.cell(row=17, column=5).value = OFZP_tr_do_rec
    tables.cell(row=18, column=5).value = SOC_OTC_tr_do_rec
    tables.cell(row=19, column=5).value = sum(Z_mat_tr_all_do_rec)
    tables.cell(row=20, column=5).value = sum(Z_zch_all_do_rec)
    tables.cell(row=21, column=5).value = Z_naklad_tr_do_rec
    tables.cell(row=22, column=5).value = OFZP_tr_do_rec + SOC_OTC_tr_do_rec + sum(Z_mat_tr_all_do_rec) +  sum(Z_zch_all_do_rec) + Z_naklad_tr_do_rec

    tables.cell(row=17, column=6).value = round((OFZP_tr_do_rec / Lg_all_do_rec) * 1000, 2)
    tables.cell(row=18, column=6).value = round((SOC_OTC_tr_do_rec / Lg_all_do_rec) * 1000, 2)
    tables.cell(row=19, column=6).value = round((sum(Z_mat_tr_all_do_rec) / Lg_all_do_rec) * 1000, 2)
    tables.cell(row=20, column=6).value = round((sum(Z_zch_all_do_rec) / Lg_all_do_rec) * 1000, 2)
    tables.cell(row=21, column=6).value = round((Z_naklad_tr_do_rec  / Lg_all_do_rec) * 1000, 2)
    tables.cell(row=22, column=6).value = round(((OFZP_tr_do_rec + SOC_OTC_tr_do_rec + sum(Z_mat_tr_all_do_rec) + sum(Z_zch_all_do_rec) + Z_naklad_tr_do_rec) / Lg_all_do_rec) * 1000, 2)

    tables.cell(row=17, column=7).value = round((OFZP_tr_do_rec / T_tr_do_rec), 2)
    tables.cell(row=18, column=7).value = round((SOC_OTC_tr_do_rec / T_tr_do_rec), 2)
    tables.cell(row=19, column=7).value = round((sum(Z_mat_tr_all_do_rec) / T_tr_do_rec), 2)
    tables.cell(row=20, column=7).value = round((sum(Z_zch_all_do_rec) / T_tr_do_rec), 2)
    tables.cell(row=21, column=7).value = round((Z_naklad_tr_do_rec / T_tr_do_rec), 2)
    tables.cell(row=22, column=7).value = round((OFZP_tr_do_rec + SOC_OTC_tr_do_rec + sum(Z_mat_tr_all_do_rec) + sum(Z_zch_all_do_rec) + Z_naklad_tr_do_rec) / T_tr_do_rec, 2)

    tables.cell(row=23, column=1).value = "table 6_2_0"

def append_tables():
    book = openpyxl.Workbook()
    tables = book.active
    table_6_0(tables)
    table_6_1(tables)
    table_6_2(tables)    
    book.save('tables.xlsx')
    book.close()


append_tables()
init_data.close()

book1 = openpyxl.Workbook()
data_output = book1.active
for i in range(len(arr_headings_output_tregub)):
    data_output.cell(row=1, column=i+1).value = arr_headings_output_tregub[i]
    data_output.cell(row=2, column=i+1).value = arr_value_output_tregub[i]
book1.save('output_treg.xlsx')
book1.close()
